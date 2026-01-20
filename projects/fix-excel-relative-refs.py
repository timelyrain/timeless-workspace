from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

wb = load_workbook('fetch-ibkr-positions-dashboard.xlsx')
ws = wb['Dashboard']

print('Fixing to relative references (no quotes, no path)...\n')

fixed_count = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.value:
            formula_text = None
            
            # Handle array formulas
            if isinstance(cell.value, ArrayFormula):
                formula_text = cell.value.text
            elif isinstance(cell.value, str) and cell.value.startswith('='):
                formula_text = cell.value
            
            if formula_text:
                # Remove any absolute path and quotes
                new_formula = formula_text.replace(
                    "'/Users/hongkiatkoh/Desktop/timeless-workspace/timeless-workspace/projects/[fetch-ibkr-positions.xlsx]PositionsHK'!",
                    "[fetch-ibkr-positions.xlsx]PositionsHK!"
                ).replace(
                    "'/Users/hongkiatkoh/Desktop/timeless-workspace/timeless-workspace/projects/[fetch-ibkr-positions.xlsx]PositionsAL'!",
                    "[fetch-ibkr-positions.xlsx]PositionsAL!"
                ).replace(
                    "'[fetch-ibkr-positions.xlsx]PositionsHK'!",
                    "[fetch-ibkr-positions.xlsx]PositionsHK!"
                ).replace(
                    "'[fetch-ibkr-positions.xlsx]PositionsAL'!",
                    "[fetch-ibkr-positions.xlsx]PositionsAL!"
                )
                
                if new_formula != formula_text:
                    cell.value = new_formula
                    print(f'{cell.coordinate}: Fixed')
                    print(f'  Before: {formula_text[:80]}')
                    print(f'  After:  {new_formula[:80]}')
                    print()
                    fixed_count += 1

wb.save('fetch-ibkr-positions-dashboard.xlsx')
print(f'✅ Fixed {fixed_count} cells')
print('\nSample corrected formula:')
print(ws['E4'].value if ws['E4'].value else 'N/A')
