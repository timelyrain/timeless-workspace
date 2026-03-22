"""
╔══════════════════════════════════════════════════════════════════════╗
║                 INSTITUTIONAL RISK SIGNAL v2.0                       ║
╚══════════════════════════════════════════════════════════════════════╝

WHAT'S NEW IN v2.0 (Jan 2026):
🎯 INSTITUTIONAL A+ GRADE - 22 indicators (was 16) with professional-quality logic
📊 CREDIT INTELLIGENCE - HY+IG spread composite, stress ratio, rate-of-change penalties
📈 OPTIONS INTELLIGENCE - VVIX, VIX9D, VXN-VIX (volatility of volatility, event risk, tech divergence)
💰 INSTITUTIONAL FLOWS - ETF flows (9 ETFs), Credit flows (4 ETFs), Sector rotation (11 sectors)
🔢 NORMALIZED SCORING - Proper /100 scale (was broken at 106/100, now 64/100)
✅ All v1.8 features preserved (2026 Portfolio Mapping, V-Recovery, Kill-Switch, Dual AI CIO)

YOUR 2026 PORTFOLIO STRUCTURE (Aligned with ARTHUR_CONTEXT.md):
- 30% Global Triads (82846, DHL, ES3, VWRA, VT, XMNE - strategic core)
- 30% Four Horsemen (CSNDX, CTEC, HEAL, INRA, GRDU - growth engine)
- 25% Cash Cow (Income Strategy: all options EXCEPT SPY/QQQ insurance)
  * Multi-leg spreads, CSPs, covered calls, iron condors, LEAPS on income stocks
  * Stock positions: SPY, QQQ, ADBE, AMD, CRM, CSCO, ORCL, COST, PEP, WMT, XOM, JPM, V, LLY, UNH, AAPL, AMZN, GOOGL, META, MSFT, NVDA, TSLA
- 2% The Alpha (Theme stocks + speculative long calls - offensive plays)
- 2.5% The Omega (Insurance: SPY/QQQ options only - bear spreads, puts, protective hedges)
- 5% The Vault (Gold - store of value)
- 5% The War Chest (Cash - dry powder)

ALLOCATION PHILOSOPHY (Institutional Risk Management):
Score 90+: FULL DEPLOYMENT - Run base allocation (2.5% base Omega)
Score 75-90: NORMAL - Base allocation, tighter stops, 1% hedging
Score 60-75: ELEVATED - Cut growth/income, raise cash to 24%, 1% hedging only
Score 40-60: HIGH RISK - Sell positions aggressively, 33% cash, 2% hedging
Score <40: EXTREME - Liquidate to 39% cash, 25% gold, 3% hedging max

HEDGING COST DISCIPLINE (Annual insurance budget: 3% of portfolio max):
- Omega allocation: 1% ELEVATED, 2% HIGH RISK, 3% EXTREME
- Estimated annual cost: $3k → $6k → $10k (stays under 3% portfolio budget)
- Primary defense: Sell positions → raise cash (War Chest) - ZERO cost
- Secondary defense: Gold (Vault) for tail risk - NO decay
- Tertiary defense: Put spreads/collars ONLY - limit theta burn
- NEVER use naked puts in elevated regimes (theta decay bankrupts portfolios)

22 INDICATORS + INSTITUTIONAL FLOWS + OPTIONS INTELLIGENCE | A+ GRADE

TIER SCORING (154 pts raw → normalized to /100):
├─ Tier 1 (45 pts): Credit composite + Fed BS + DXY
├─ Tier 2 (60 pts): Positioning + Institutional flows (ETF/Credit/Sector)
└─ Tier 3 (49 pts): Options intelligence (VIX Term, SKEW, VVIX, VIX9D, VXN) + Structure

TODO - FUTURE ENHANCEMENTS (Requires Paid Data):
❌ CBOE Put/Call Ratio - yfinance options data too unreliable/incomplete
   Recommended providers:
   • Polygon.io ($199/mo) - Best API quality, institutional-grade
   • Barchart ($30/mo) - Most affordable with CBOE P/C ratio access
   • Alpha Vantage Premium ($49/mo) - Good balance of cost/quality
   Current workaround: VIX/VVIX/SKEW/VIX9D cover sentiment adequately

✅ VALIDATION FRAMEWORK (No Additional Cost):
   Walk-forward validation using existing risk_history.json data
   • Run projects/instituitional-risk-signal-validation-report-monthly.py monthly to validate signal quality
   • Analyzes score correlation vs SPY forward returns
   • Checks regime change accuracy vs actual drawdowns
   • Alert timing analysis (did warnings precede drops?)
   • No extra API calls - uses data already collected daily

SYSTEM: 15% Portfolio Risk Monitor (Arthur Protocol)
CONTEXT: See ARTHUR_CONTEXT.md for full strategy and logic constraints.
"""

import yfinance as yf
import pandas as pd
from fredapi import Fred
import requests
from datetime import datetime, timedelta
import os
from pathlib import Path
from dotenv import load_dotenv
import json
from portfolio_categories_mappings import SYMBOL_MAPPING

# =============================================================================
# CONFIGURATION
# =============================================================================

env_path = Path(__file__).parent / '.env'
load_dotenv(env_path)

CONFIG = {
    'FRED_API_KEY': os.getenv('FRED_API_KEY', 'YOUR_FRED_API_KEY_HERE'),
    'TELEGRAM_TOKEN': os.getenv('TELEGRAM_TOKEN_RISK'),
    'CHAT_ID': os.getenv('CHAT_ID'),
    'RUN_TIME': '09:15',
    
    # V-Recovery Override Settings
    'V_RECOVERY_ENABLED': True,
    'V_RECOVERY_SCORE_THRESHOLD': 30,
    'V_RECOVERY_SPY_GAIN': 8,
    'V_RECOVERY_SPY_DAYS': 10,
    'V_RECOVERY_VIX_DROP': 15,
    'V_RECOVERY_LOOKBACK': 30,
}

# Portfolio Configuration (2026 Structure - Aligned with ARTHUR_CONTEXT.md)
PORTFOLIO_2026 = {
    'TOTAL_CAPITAL': 1_000_000,  # $1M active trading capital
    'BASE_ALLOCATION': {
        'global_triads': 0.30,      # 82846, DHL, ES3, VWRA, VT, XMNE
        'four_horsemen': 0.30,      # CSNDX, CTEC, HEAL, INRA, GRDU
        'cash_cow': 0.25,           # Wheel on GOOGL, PEP, V
        'alpha': 0.02,              # Theme stocks + call options
        'omega': 0.025,             # QQQ puts + bear spreads
        'vault': 0.05,              # Gold (GLD/IAU)
        'war_chest': 0.05           # Cash holdings
    }
}

if CONFIG['FRED_API_KEY'] == 'YOUR_FRED_API_KEY_HERE':
    print("⚠️  WARNING: FRED_API_KEY not set.")
    print("   Get your free key at: https://fred.stlouisfed.org/docs/api/api_key.html")

fred = Fred(api_key=CONFIG['FRED_API_KEY'])

# =============================================================================
# TELEGRAM UTILITIES
# =============================================================================

def send_to_telegram(message):
    """Send message to Telegram channel"""
    if not CONFIG['TELEGRAM_TOKEN'] or not CONFIG['CHAT_ID']:
        print("⚠️  Telegram credentials not found. Skipping alert.")
        return False
    
    url = f"https://api.telegram.org/bot{CONFIG['TELEGRAM_TOKEN']}/sendMessage"
    max_length = 4000
    messages = []
    
    if len(message) <= max_length:
        messages = [message]
    else:
        parts = message.split('\n\n')
        current = ""
        for part in parts:
            if len(current) + len(part) + 2 <= max_length:
                current += part + "\n\n"
            else:
                if current:
                    messages.append(current.strip())
                current = part + "\n\n"
        if current:
            messages.append(current.strip())
    
    success = True
    for i, msg in enumerate(messages):
        payload = {'chat_id': CONFIG['CHAT_ID'], 'text': msg, 'disable_web_page_preview': True}
        try:
            resp = requests.post(url, json=payload, timeout=10)
            if resp.status_code == 200:
                print(f"✅ Sent to Telegram" + (f" (part {i+1}/{len(messages)})" if len(messages) > 1 else ""))
            else:
                print(f"⚠️  Telegram error: {resp.status_code}")
                success = False
        except Exception as e:
            print(f"⚠️  Telegram send failed: {e}")
            success = False
    
    return success

# =============================================================================
# HISTORICAL DATA MANAGER
# =============================================================================

class HistoricalDataManager:
    """Manages historical risk scores and market data"""
    
    def __init__(self, history_file='risk_history.json'):
        self.history_file = history_file
        self.history = self._load_history()
    
    def _load_history(self):
        if os.path.exists(self.history_file):
            try:
                with open(self.history_file, 'r') as f:
                    return json.load(f)
            except:
                return {'scores': [], 'overrides': [], 'backwardation': [], 'alerts': []}
        return {'scores': [], 'overrides': [], 'backwardation': [], 'alerts': []}
    
    def save_history(self):
        """Save risk history to JSON file"""
        try:
            with open(self.history_file, 'w') as f:
                json.dump(self.history, f, indent=2)
        except Exception as e:
            print(f"⚠️  Could not save history: {e}")
    
    def add_score(self, date, score, spy_price, vix, vix_structure=None, vixy_vxx_ratio=None):
        """Add daily risk score entry to history and maintain 90-day rolling window"""
        entry = {'date': date, 'score': score, 'spy': spy_price, 'vix': vix}
        if vix_structure:
            entry['vix_structure'] = vix_structure
        if vixy_vxx_ratio:
            entry['vixy_vxx_ratio'] = vixy_vxx_ratio
        
        self.history['scores'].append(entry)
        cutoff_date = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
        self.history['scores'] = [s for s in self.history['scores'] if s['date'] >= cutoff_date]
    
    def add_backwardation_event(self, date, vixy_vxx_ratio, magnitude_pct):
        """Record VIX backwardation event for tracking consecutive occurrences"""
        if 'backwardation' not in self.history:
            self.history['backwardation'] = []
        self.history['backwardation'].append({'date': date, 'ratio': vixy_vxx_ratio, 'magnitude': magnitude_pct})
        cutoff_date = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
        self.history['backwardation'] = [b for b in self.history['backwardation'] if b['date'] >= cutoff_date]
    
    def get_backwardation_streak(self):
        """Count consecutive days of VIX backwardation (for extreme fear detection)"""
        if 'backwardation' not in self.history or not self.history['backwardation']:
            return 0, 0.0
        events = sorted(self.history['backwardation'], key=lambda x: x['date'], reverse=True)
        today = datetime.now().strftime('%Y-%m-%d')
        streak = 0
        magnitudes = []
        if events and events[0]['date'] == today:
            streak = 1
            magnitudes.append(events[0]['magnitude'])
            for i in range(1, len(events)):
                prev_date = (datetime.strptime(events[i-1]['date'], '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
                if events[i]['date'] == prev_date:
                    streak += 1
                    magnitudes.append(events[i]['magnitude'])
                else:
                    break
        avg_magnitude = sum(magnitudes) / len(magnitudes) if magnitudes else 0.0
        return streak, avg_magnitude
    
    def add_override_event(self, date, reason, conditions):
        """Record V-Recovery override event with reason and triggering conditions"""
        self.history['overrides'].append({'date': date, 'reason': reason, 'conditions': conditions})
        if len(self.history['overrides']) > 50:
            self.history['overrides'] = self.history['overrides'][-50:]
    
    def get_override_streak(self):
        """Count consecutive days of override triggers ending yesterday
        Used for the Kill-Switch logic
        """
        if 'overrides' not in self.history or not self.history['overrides']:
            return 0
            
        # Sort by date descending
        events = sorted(self.history['overrides'], key=lambda x: x['date'], reverse=True)
        streak = 0
        
        # Check from yesterday backwards (to see how long we've been in override mode)
        check_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        
        for event in events:
            if event['date'] == check_date:
                streak += 1
                # Move check_date back one day
                check_date = (datetime.strptime(check_date, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
            elif event['date'] > check_date:
                continue  # Skip if duplicate or future (unlikely)
            else:
                break  # Gap found
                
        return streak
    
    def get_recent_scores(self, days=30):
        """Get risk scores from the last N days"""
        cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        return [s for s in self.history['scores'] if s['date'] >= cutoff]
    
    def had_extreme_score_recently(self, threshold=30, days=30):
        """Check if risk score fell below threshold within specified days (for kill-switch)"""
        recent = self.get_recent_scores(days)
        if not recent:
            return False
        return min(s['score'] for s in recent) < threshold
    
    def get_override_start_date(self):
        """Get the date when current override started"""
        if not self.history.get('overrides'):
            return None
        # Get most recent override
        recent_overrides = sorted(self.history['overrides'], key=lambda x: x['date'], reverse=True)
        if recent_overrides:
            return recent_overrides[0]['date']
        return None
    
    def add_drift_snapshot(self, date, total_drift, category_drifts):
        """Store daily portfolio drift snapshot"""
        if 'drift_history' not in self.history:
            self.history['drift_history'] = []
        
        entry = {
            'date': date,
            'total_drift': total_drift,
            'categories': category_drifts
        }
        
        # Remove existing entry for same date (in case of re-run)
        self.history['drift_history'] = [d for d in self.history['drift_history'] if d['date'] != date]
        self.history['drift_history'].append(entry)
        
        # Keep last 90 days
        cutoff_date = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
        self.history['drift_history'] = [d for d in self.history['drift_history'] if d['date'] >= cutoff_date]
    
    def get_drift_trend(self, days=7):
        """Get drift trend over last N days"""
        if 'drift_history' not in self.history or not self.history['drift_history']:
            return None
        
        cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        recent = sorted(
            [d for d in self.history['drift_history'] if d['date'] >= cutoff],
            key=lambda x: x['date']
        )
        
        if len(recent) < 2:
            return None
        
        # Calculate trend (improving = drift decreasing)
        first_drift = recent[0]['total_drift']
        last_drift = recent[-1]['total_drift']
        change = last_drift - first_drift
        change_pct = (change / first_drift * 100) if first_drift > 0 else 0
        
        return {
            'history': recent,
            'first_drift': first_drift,
            'last_drift': last_drift,
            'change': change,
            'change_pct': change_pct,
            'improving': change < 0,  # Drift decreasing = improving
            'days': len(recent)
        }
    
    def add_indicator_data(self, date, indicators):
        """Store daily indicator values for historical comparison"""
        if 'indicators' not in self.history:
            self.history['indicators'] = []
        
        # Remove existing entry for same date
        self.history['indicators'] = [d for d in self.history['indicators'] if d.get('date') != date]
        
        entry = {'date': date, **indicators}
        self.history['indicators'].append(entry)
        
        # Keep last 90 days
        cutoff_date = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
        self.history['indicators'] = [d for d in self.history['indicators'] if d.get('date', '') >= cutoff_date]
    
    def get_spread_change(self, indicator_name, days=30):
        """Calculate N-day rate of change for credit spreads"""
        if not self.history.get('scores'):
            return None
        
        cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        recent = sorted(
            [s for s in self.history['scores'] if s['date'] >= cutoff and indicator_name in s.get('data', {})],
            key=lambda x: x['date']
        )
        
        if len(recent) < 2:
            return None
        
        old_value = recent[0]['data'][indicator_name]
        new_value = recent[-1]['data'][indicator_name]
        
        if old_value is None or new_value is None or old_value == 0:
            return None
        
        change_pct = ((new_value - old_value) / old_value) * 100
        return {
            'old_value': old_value,
            'new_value': new_value,
            'change_pct': change_pct,
            'days': len(recent),
            'widening': change_pct > 0
        }
    
    def get_score_trend(self, days=7):
        """Get risk score trend with regime changes"""
        if not self.history.get('scores'):
            return None
        
        cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        recent = sorted(
            [s for s in self.history['scores'] if s['date'] >= cutoff],
            key=lambda x: x['date']
        )
        
        if len(recent) < 2:
            return None
        
        # Calculate score change
        first_score = recent[0]['score']
        last_score = recent[-1]['score']
        change = last_score - first_score
        
        # Map scores to regimes
        def get_regime(score):
            """Convert numeric risk score to regime label"""
            if score >= 90: return 'ALL CLEAR'
            elif score >= 75: return 'NORMAL'
            elif score >= 60: return 'ELEVATED'
            elif score >= 40: return 'HIGH RISK'
            else: return 'EXTREME RISK'
        
        first_regime = get_regime(first_score)
        last_regime = get_regime(last_score)
        regime_changed = first_regime != last_regime
        
        # Track key signal changes
        signal_changes = {}
        if 'vix' in recent[0] and 'vix' in recent[-1]:
            vix_change = recent[-1]['vix'] - recent[0]['vix']
            vix_pct = (vix_change / recent[0]['vix'] * 100) if recent[0]['vix'] > 0 else 0
            signal_changes['vix'] = {'change': vix_change, 'pct': vix_pct, 'first': recent[0]['vix'], 'last': recent[-1]['vix']}
        
        return {
            'history': recent,
            'first_score': first_score,
            'last_score': last_score,
            'change': change,
            'improving': change > 0,  # Score increasing = improving
            'first_regime': first_regime,
            'last_regime': last_regime,
            'regime_changed': regime_changed,
            'signal_changes': signal_changes,
            'days': len(recent)
        }
    
    def should_suppress_alert(self, alert_type, days=1):
        """Check if alert was already sent in the last N days (prevent spam)"""
        if 'alerts' not in self.history:
            self.history['alerts'] = []
        
        cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        recent_alerts = [a for a in self.history['alerts'] if a.get('date', '') >= cutoff]
        
        # Check if this alert type was sent recently
        return any(a.get('type') == alert_type for a in recent_alerts)
    
    def add_alert(self, date, alert_type, message):
        """Record sent alert for deduplication"""
        if 'alerts' not in self.history:
            self.history['alerts'] = []
        
        self.history['alerts'].append({
            'date': date,
            'type': alert_type,
            'message': message
        })
        
        # Keep only last 30 days of alerts
        cutoff_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        self.history['alerts'] = [a for a in self.history['alerts'] if a.get('date', '') >= cutoff_date]
    
    def days_since_override_start(self):
        """Calculate days since override activated"""
        start_date = self.get_override_start_date()
        if not start_date:
            return 0
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            return (datetime.now() - start).days
        except:
            return 0

# =============================================================================
# RISK DASHBOARD - v2.0 with 22 Institutional Indicators
# =============================================================================

class RiskDashboard:
    # Symbol to category mapping for IBKR positions (from fetch-ibkr-positions-dashboard.xlsx)
    # NOTE: Stock positions categorized by underlying ticker. Options categorization logic:
    # - Cash Cow: All options EXCEPT SPY/QQQ (spreads, CSPs, CCs, iron condors, LEAPS)
    # - Omega: SPY/QQQ options ONLY (bear spreads, puts, protective hedges)
    # - Alpha: Long calls on non-income stocks (speculative)
    # SYMBOL_MAPPING imported from config.py
    
    def __init__(self):
        self.data = {}
        self.scores = {}
        self.alerts = []
        self.timestamp = datetime.now()
        self.history_manager = HistoricalDataManager()
        self.v_recovery_active = False
        self.v_recovery_reason = None
        self.missing_signals = []
        self.actual_positions = None
    
    # [DATA FETCHING METHODS - Enhanced in v2.0 with institutional flows]
    # v2.1: Upgraded from S&P 100 (98 stocks) to S&P 500 (503 stocks) for breadth analysis
    #        Uses batch yf.download() instead of individual ticker calls (8s vs 5min)
    
    def _get_sp500_tickers(self):
        """Get S&P 500 tickers for breadth analysis.
        
        Fetches from GitHub-hosted CSV (datasets/s-and-p-500-companies) with
        local cache (sp500_cache.json, refreshed every 30 days).
        Falls back to hardcoded list of ~500 tickers if fetch fails.
        
        Returns: list of ticker strings (e.g., ['AAPL', 'MSFT', ...])
        """
        import pandas as pd
        from pathlib import Path
        cache_file = Path(__file__).parent / 'sp500_cache.json'
        
        # Check if cache exists and is fresh (<30 days)
        try:
            if os.path.exists(cache_file):
                with open(cache_file, 'r') as f:
                    cache = json.load(f)
                cache_date = cache.get('timestamp', '')
                tickers = cache.get('tickers', [])
                if tickers and len(tickers) >= 450:
                    age_days = (datetime.now() - datetime.strptime(cache_date, '%Y-%m-%d')).days
                    if age_days <= 30:
                        return tickers
                    else:
                        print(f"   ⚠️  sp500_cache.json is {age_days} days old, refreshing...")
        except Exception:
            pass
        
        # Fetch fresh list from GitHub datasets
        try:
            url = 'https://raw.githubusercontent.com/datasets/s-and-p-500-companies/main/data/constituents.csv'
            sp500 = pd.read_csv(url)
            tickers = sorted(sp500['Symbol'].str.replace('.', '-', regex=False).tolist())
            if len(tickers) >= 450:
                cache_data = {
                    'timestamp': datetime.now().strftime('%Y-%m-%d'),
                    'source': 'github/datasets/s-and-p-500-companies',
                    'count': len(tickers),
                    'tickers': tickers,
                }
                with open(cache_file, 'w') as f:
                    json.dump(cache_data, f, indent=2)
                print(f"   ✅ Refreshed sp500_cache.json: {len(tickers)} tickers")
                return tickers
        except Exception as e:
            print(f"   ⚠️  GitHub S&P 500 fetch failed: {e}, using fallback")
        
        # Try loading stale cache (better than hardcoded)
        try:
            with open(cache_file, 'r') as f:
                cache = json.load(f)
                tickers = cache.get('tickers', [])
                if tickers and len(tickers) >= 450:
                    print(f"   ⚠️  Using stale sp500_cache.json ({len(tickers)} tickers)")
                    return tickers
        except Exception:
            pass
        
        # Hardcoded fallback — S&P 500 constituents (updated March 2026)
        print(f"   ⚠️  Using hardcoded S&P 500 fallback list")
        return [
            'A','AAPL','ABBV','ABNB','ABT','ACGL','ACN','ADBE','ADI','ADM',
            'ADP','ADSK','AEE','AEP','AES','AFL','AIG','AIZ','AJG','AKAM',
            'ALB','ALGN','ALL','ALLE','AMAT','AMCR','AMD','AME','AMGN','AMP',
            'AMT','AMZN','ANET','AON','AOS','APA','APD','APH','APO','APP',
            'APTV','ARE','ARES','ATO','AVB','AVGO','AVY','AWK','AXON','AXP',
            'AZO','BA','BAC','BALL','BAX','BBY','BDX','BEN','BF-B','BG',
            'BIIB','BK','BKNG','BKR','BLDR','BLK','BMY','BR','BRK-B','BRO',
            'BSX','BX','BXP','C','CAG','CAH','CARR','CAT','CB','CBOE',
            'CBRE','CCI','CCL','CDNS','CDW','CEG','CF','CFG','CHD','CHRW',
            'CHTR','CI','CINF','CL','CLX','CMCSA','CME','CMG','CMI','CMS',
            'CNC','CNP','COF','COO','COP','COR','COST','CPAY','CPB','CPRT',
            'CPT','CRL','CRM','CRWD','CSCO','CSGP','CSX','CTAS','CTRA','CTSH',
            'CTVA','CVS','CVX','D','DAL','DD','DE','DELL','DG','DGX',
            'DHI','DHR','DIS','DLR','DLTR','DOV','DOW','DPZ','DRI','DTE',
            'DUK','DVA','DVN','DXCM','EA','EBAY','ECL','ED','EFX','EIX',
            'EL','ELV','EME','EMR','EOG','EPAM','EQIX','EQR','EQT','ERIE',
            'ES','ESS','ETN','ETR','EVRG','EW','EXC','EXPD','EXPE','EXR',
            'F','FANG','FAST','FCX','FDS','FDX','FE','FFIV','FICO','FIS',
            'FISV','FITB','FOX','FOXA','FRT','FSLR','FTNT','FTV','GD','GDDY',
            'GE','GEHC','GEN','GEV','GILD','GIS','GL','GLW','GM','GNRC',
            'GOOG','GOOGL','GPC','GPN','GRMN','GS','GWW','HAL','HAS','HBAN',
            'HCA','HD','HIG','HII','HLT','HOLX','HON','HPE','HPQ','HRL',
            'HSIC','HST','HSY','HUBB','HUM','HWM','IBM','ICE','IDXX','IEX',
            'IFF','INCY','INTC','INTU','INVH','IP','IQV','IR','IRM','ISRG',
            'IT','ITW','IVZ','J','JBHT','JBL','JCI','JKHY','JNJ','JPM',
            'KDP','KEY','KEYS','KHC','KIM','KKR','KLAC','KMB','KMI','KO',
            'KR','KVUE','L','LDOS','LEN','LH','LHX','LIN','LLY','LMT',
            'LNT','LOW','LRCX','LULU','LUV','LVS','LW','LYB','LYV','MA',
            'MAA','MAR','MAS','MCD','MCHP','MCK','MCO','MDLZ','MDT','MET',
            'META','MGM','MKC','MLM','MMM','MNST','MO','MOH','MOS','MPC',
            'MPWR','MRK','MRNA','MS','MSCI','MSFT','MSI','MTB','MTCH','MTD',
            'MU','NCLH','NDAQ','NDSN','NEE','NEM','NFLX','NI','NKE','NOC',
            'NOW','NRG','NSC','NTAP','NTRS','NUE','NVDA','NVR','NWS','NWSA',
            'NXPI','O','ODFL','OKE','OMC','ON','ORCL','ORLY','OTIS','OXY',
            'PANW','PAYC','PAYX','PCAR','PCG','PEG','PEP','PFE','PFG','PG',
            'PGR','PH','PHM','PKG','PLD','PLTR','PM','PNC','PNR','PNW',
            'POOL','PPG','PPL','PRU','PSA','PSX','PTC','PWR','PYPL','QCOM',
            'RCL','REG','REGN','RF','RJF','RL','RMD','ROK','ROL','ROP',
            'ROST','RSG','RTX','RVTY','SBAC','SBUX','SCHW','SHW','SJM','SLB',
            'SMCI','SNA','SNPS','SO','SPG','SPGI','SRE','STE','STLD','STT',
            'STX','STZ','SWK','SWKS','SYF','SYK','SYY','T','TAP','TDG',
            'TDY','TECH','TEL','TER','TFC','TGT','TJX','TMO','TMUS','TPR',
            'TRGP','TRMB','TROW','TRV','TSCO','TSLA','TSN','TT','TTWO','TXN',
            'TXT','TYL','UAL','UBER','UDR','UHS','ULTA','UNH','UNP','UPS',
            'URI','USB','V','VICI','VLO','VLTO','VMC','VRSK','VRSN','VRTX',
            'VTR','VTRS','VZ','WAB','WAT','WBD','WDAY','WDC','WEC','WELL',
            'WFC','WM','WMB','WMT','WRB','WST','WTW','WY','WYNN','XEL',
            'XOM','XYL','YUM','ZBH','ZBRA','ZTS',
        ]
    
    def _batch_breadth_scan(self):
        """Batch-download S&P 500 data and compute all breadth indicators at once.
        
        Returns dict with:
          - pct_above_50ma: float (percentage of stocks above 50-day MA)
          - pct_below_200ma: float (percentage of stocks below 200-day MA)
          - new_hl: int (new highs minus new lows over 3 months)
          - breadth_valid_50: int (count of valid stocks for 50MA)
          - breadth_valid_200: int (count of valid stocks for 200MA)
          - breadth_valid_hl: int (count of valid stocks for H-L)
        
        Uses yf.download() batch mode — ~8 seconds for 500 tickers vs 5+ min individual.
        """
        tickers = self._get_sp500_tickers()
        result = {}
        
        try:
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                data = yf.download(tickers, period='1y', progress=False, threads=True)
            
            close = data['Close']
            latest = close.iloc[-1]
            
            # % Above 50-day MA
            ma50 = close.rolling(50).mean().iloc[-1]
            mask_50 = latest.notna() & ma50.notna()
            valid_50 = int(mask_50.sum())
            above_50 = int(((latest > ma50) & mask_50).sum())
            if valid_50 > 0:
                pct = above_50 / valid_50 * 100
                result['pct_above_50ma'] = pct
                result['breadth_valid_50'] = valid_50
                print(f"   ✓ % Above 50-MA: {pct:.0f}% ({above_50}/{valid_50} S&P 500)")
            
            # % Below 200-day MA
            ma200 = close.rolling(200).mean().iloc[-1]
            mask_200 = latest.notna() & ma200.notna()
            valid_200 = int(mask_200.sum())
            below_200 = int(((latest < ma200) & mask_200).sum())
            if valid_200 > 0:
                pct = below_200 / valid_200 * 100
                result['pct_below_200ma'] = pct
                result['breadth_valid_200'] = valid_200
                print(f"   ✓ % Below 200-MA: {pct:.0f}% ({below_200}/{valid_200} S&P 500)")
            
            # New Highs minus Lows (3-month window)
            close_3m = close.iloc[-63:]  # ~3 months of trading days
            if len(close_3m) >= 52:
                cur = close_3m.iloc[-1]
                high_3m = close_3m.max()
                low_3m = close_3m.min()
                # Within 0.5% of 3-month high = new high; within 0.5% of 3-month low = new low
                mask_hl = cur.notna() & high_3m.notna() & low_3m.notna()
                highs = int(((cur >= high_3m * 0.995) & mask_hl).sum())
                lows = int(((cur <= low_3m * 1.005) & mask_hl).sum())
                valid_hl = int(mask_hl.sum())
                net = highs - lows
                result['new_hl'] = net
                result['breadth_valid_hl'] = valid_hl
                print(f"   ✓ New H-L: {net:+d} (H:{highs} L:{lows} of {valid_hl} S&P 500)")
            
            # Store close data for A/D Line computation (avoids re-downloading)
            result['_batch_close'] = close
            
        except Exception as e:
            print(f"   ✗ Batch breadth scan failed: {e}")
            self.missing_signals.extend(['% Above 50MA', '% Below 200MA', 'New H-L'])
        
        return result
    
    def _fred_get(self, series, name):
        try:
            data = fred.get_series_latest_release(series)
            val = float(data.iloc[-1])
            print(f"   ✓ {name}: {val:.2f}")
            return val
        except Exception as e:
            print(f"   ✗ {name}: Error")
            self.missing_signals.append(name)
            return None
    
    def _yf_get(self, ticker, name):
        try:
            val = yf.Ticker(ticker).history(period='1d')['Close'].iloc[-1]
            print(f"   ✓ {name}: {val:.2f}")
            return float(val)
        except:
            print(f"   ✗ {name}: Error")
            self.missing_signals.append(name)
            return None
    
    def _fed_bs_yoy(self):
        try:
            bs = fred.get_series_latest_release('WALCL')
            yoy = ((bs.iloc[-1] - bs.iloc[-52]) / bs.iloc[-52]) * 100
            print(f"   ✓ Fed BS YoY: {yoy:.1f}%")
            return float(yoy)
        except:
            print(f"   ✗ Fed BS YoY: Error")
            self.missing_signals.append("Fed BS YoY")
            return None
    
    def _dxy_trend(self):
        try:
            hist = yf.Ticker('DX-Y.NYB').history(period='2mo')['Close']
            if len(hist) < 20:
                return None
            trend = ((hist.iloc[-1] - hist.rolling(20).mean().iloc[-1]) / hist.rolling(20).mean().iloc[-1]) * 100
            print(f"   ✓ DXY Trend: {trend:.1f}%")
            return float(trend)
        except:
            print(f"   ✗ DXY Trend: Error")
            return None
    
    def _pct_above_ma(self, period):
        """DEPRECATED: Use _batch_breadth_scan() instead. Kept as fallback."""
        return None
    
    def _pct_below_ma(self, period):
        """DEPRECATED: Use _batch_breadth_scan() instead. Kept as fallback."""
        return None
    
    def _ad_line_status(self, batch_close=None):
        """Real Advance/Decline breadth from S&P 500 batch data.
        
        Computes daily advance/decline ratio over recent days and compares
        to 20-day average to detect breadth divergence.
        
        Uses the batch S&P 500 close data already downloaded by _batch_breadth_scan()
        to avoid redundant API calls.
        
        Returns: float — A/D ratio (advances/total). Range 0.0-1.0.
          >0.55 = Confirming (broad participation)
          0.45-0.55 = Flat (mixed)
          <0.45 = Diverging (narrow/weak breadth)
        """
        try:
            if batch_close is None or batch_close.empty:
                print(f"   ✗ AD Line: No batch data")
                return None
            
            close = batch_close.dropna(axis=1, how='all')
            if close.shape[0] < 25 or close.shape[1] < 100:
                print(f"   ✗ AD Line: Insufficient data ({close.shape[1]} tickers)")
                return None
            
            # Daily advances: stocks that closed higher than previous day
            daily_returns = close.pct_change(fill_method=None)
            advances = (daily_returns > 0).sum(axis=1)
            declines = (daily_returns < 0).sum(axis=1)
            total = advances + declines
            
            # Avoid division by zero
            ad_ratio = advances / total.replace(0, float('nan'))
            ad_ratio = ad_ratio.dropna()
            
            if len(ad_ratio) < 20:
                print(f"   ✗ AD Line: Insufficient history")
                return None
            
            # 5-day average A/D ratio (recent breadth)
            recent_ad = float(ad_ratio.iloc[-5:].mean())
            # 20-day average (baseline)
            baseline_ad = float(ad_ratio.iloc[-20:].mean())
            
            # Latest day advances/declines for display
            latest_adv = int(advances.iloc[-1])
            latest_dec = int(declines.iloc[-1])
            
            print(f"   ✓ AD Breadth: {recent_ad:.1%} adv rate (5d), baseline {baseline_ad:.1%} (20d) | Today: {latest_adv}A/{latest_dec}D")
            return recent_ad
        except Exception as e:
            print(f"   ✗ AD Line: Error - {e}")
            return None
    
    def _new_highs_lows(self):
        """DEPRECATED: Use _batch_breadth_scan() instead. Kept as fallback."""
        return None
    
    def _sector_rotation(self):
        """Sector Rotation: Defensive vs Cyclical basket ratio trend
        Cyclicals: XLY (Consumer Disc), XLF (Financials), XLI (Industrials)
        Defensives: XLU (Utilities), XLP (Consumer Staples), XLV (Healthcare)
        
        Positive trend = defensives outperforming (risk-off)
        Negative trend = cyclicals outperforming (risk-on)
        
        Institutional standard: basket approach eliminates single-sector noise
        (e.g., tech selling off on regulation doesn't false-flag as risk-off)
        """
        try:
            import pandas as pd
            cyclicals = ['XLY', 'XLF', 'XLI']
            defensives = ['XLU', 'XLP', 'XLV']
            all_tickers = cyclicals + defensives
            
            data = yf.download(all_tickers, period='2mo', progress=False, threads=True)
            if data.empty:
                return None
            
            close = data['Close']
            
            # Verify sufficient data for all tickers
            for t in all_tickers:
                if t not in close.columns or close[t].dropna().shape[0] < 20:
                    print(f"   ✗ Sector Rotation: Insufficient data for {t}")
                    return None
            
            # Equal-weight basket: normalize each to day-0 = 1, then average
            # This prevents large-cap ETFs from dominating the basket
            cyc_norm = close[cyclicals].div(close[cyclicals].iloc[0])
            def_norm = close[defensives].div(close[defensives].iloc[0])
            
            cyc_basket = cyc_norm.mean(axis=1)
            def_basket = def_norm.mean(axis=1)
            
            # Defensive/Cyclical ratio — rising = defensives leading (risk-off)
            ratio = def_basket / cyc_basket
            ratio = ratio.dropna()
            
            if len(ratio) < 20:
                return None
            
            ma20 = ratio.rolling(20).mean()
            current = ratio.iloc[-1]
            ma20_val = ma20.iloc[-1]
            
            if pd.isna(current) or pd.isna(ma20_val) or ma20_val == 0:
                return None
            
            trend = ((current - ma20_val) / ma20_val) * 100
            print(f"   ✓ Def/Cyc Basket: {trend:+.1f}% (XLU+XLP+XLV vs XLY+XLF+XLI)")
            return float(trend)
        except Exception as e:
            print(f"   ✗ Sector Rotation: Error - {e}")
            return None
    
    def _gold_spy_ratio(self):
        try:
            import pandas as pd
            gld = yf.Ticker('GLD').history(period='2mo')['Close']
            spy = yf.Ticker('SPY').history(period='2mo')['Close']
            if len(gld) < 20 or len(spy) < 20:
                return None
            ratio = gld / spy
            current = ratio.iloc[-1]
            ma20 = ratio.rolling(20).mean().iloc[-1]
            
            # Check for NaN values before calculation
            if pd.isna(current) or pd.isna(ma20) or ma20 == 0:
                print(f"   ✗ GLD/SPY: Invalid data")
                return None
            
            trend = ((current - ma20) / ma20) * 100
            print(f"   ✓ GLD/SPY: {trend:+.1f}%")
            return float(trend)
        except Exception as e:
            print(f"   ✗ GLD/SPY: Error")
            return None
    
    def _vix_term_structure(self):
        """VIX Term Structure: Measures options market sentiment via VIX futures curve
        Contango (VIX3M > VIX) = Complacency, Backwardation = Fear
        """
        try:
            vix_spot = yf.Ticker('^VIX').history(period='5d')['Close'].iloc[-1]
            vix_3m = yf.Ticker('^VIX3M').history(period='5d')['Close'].iloc[-1]
            
            # Calculate term structure slope (3M vs Spot)
            term_slope_pct = ((vix_3m - vix_spot) / vix_spot) * 100
            
            # Determine structure
            if term_slope_pct > 25:
                struct = 'Extreme Contango'
            elif term_slope_pct > 15:
                struct = 'Contango'
            elif term_slope_pct > -5:
                struct = 'Flat'
            else:
                struct = 'Backwardation'
            
            print(f"   ✓ VIX Term: {struct} ({term_slope_pct:+.1f}%)")
            return term_slope_pct, struct
        except Exception as e:
            print(f"   ✗ VIX Term: Error")
            return None, None
    
    def _skew_index(self):
        """CBOE SKEW Index: Measures tail risk in options market
        >140 = Crash fear, 100-130 = Normal, <100 = Complacency
        """
        try:
            skew_data = yf.Ticker('^SKEW').history(period='5d')
            if len(skew_data) > 0:
                skew = skew_data['Close'].iloc[-1]
                print(f"   ✓ SKEW: {skew:.1f}")
                return float(skew)
            else:
                print(f"   ✗ SKEW: No data")
                return None
        except Exception as e:
            print(f"   ✗ SKEW: Error")
            return None
    
    def _vvix_index(self):
        """VVIX (VIX of VIX): Measures uncertainty about volatility
        Spikes before major market moves, leading indicator for VIX itself
        """
        try:
            vvix_data = yf.Ticker('^VVIX').history(period='5d')
            if len(vvix_data) > 0:
                vvix = vvix_data['Close'].iloc[-1]
                print(f"   ✓ VVIX: {vvix:.1f}")
                return float(vvix)
            else:
                print(f"   ✗ VVIX: No data")
                return None
        except Exception as e:
            print(f"   ✗ VVIX: Error")
            return None
    
    def _vix9d_ratio(self):
        """VIX9D/VIX Ratio: Short-term (9-day) vs medium-term (30-day) fear
        VIX9D > VIX = Elevated near-term event risk
        VIX9D < VIX = Near-term calm but medium-term concern
        """
        try:
            vix9d_data = yf.Ticker('^VIX9D').history(period='5d')
            vix_data = yf.Ticker('^VIX').history(period='5d')
            
            if len(vix9d_data) > 0 and len(vix_data) > 0:
                vix9d = vix9d_data['Close'].iloc[-1]
                vix = vix_data['Close'].iloc[-1]
                
                if vix > 0:
                    ratio_pct = ((vix9d / vix) - 1) * 100
                    print(f"   ✓ VIX9D/VIX: {ratio_pct:+.1f}%")
                    return float(ratio_pct)
                else:
                    print(f"   ✗ VIX9D/VIX: Invalid VIX")
                    return None
            else:
                print(f"   ✗ VIX9D/VIX: No data")
                return None
        except Exception as e:
            print(f"   ✗ VIX9D/VIX: Error")
            return None
    
    def _vxn_vix_spread(self):
        """VXN-VIX Spread: NASDAQ-100 vol vs S&P 500 vol
        Positive spread = Tech fear > Market fear (QQQ risk)
        Negative spread = Tech calm, broader market fear
        """
        try:
            vxn_data = yf.Ticker('^VXN').history(period='5d')
            vix_data = yf.Ticker('^VIX').history(period='5d')
            
            if len(vxn_data) > 0 and len(vix_data) > 0:
                vxn = vxn_data['Close'].iloc[-1]
                vix = vix_data['Close'].iloc[-1]
                spread = vxn - vix
                print(f"   ✓ VXN-VIX: {spread:+.2f}")
                return float(spread)
            else:
                print(f"   ✗ VXN-VIX: No data")
                return None
        except Exception as e:
            print(f"   ✗ VXN-VIX: Error")
            return None
    
    def _etf_flow_divergence(self):
        """ETF Flow Divergence: Institutional money flow across 9 major ETFs
        Tracks volume surge + price momentum to detect inflows/outflows.
        Uses batch yf.download() for efficiency.
        
        Risk-on ETFs: SPY, QQQ, IWM, DIA, EEM (equity)
        Risk-off ETFs: TLT, GLD (safe haven)
        Credit ETFs: HYG, LQD (credit health)
        
        Returns tuple: (inflow_count, outflow_count, net_flow_score)
        net_flow_score: weighted sum of per-ETF flow signals [-1 to +1 each]
        """
        try:
            etfs = ['SPY', 'QQQ', 'IWM', 'DIA', 'EEM', 'TLT', 'GLD', 'HYG', 'LQD']
            
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                data = yf.download(etfs, period='2mo', progress=False, threads=True)
            
            if data.empty:
                print(f"   ✗ ETF Flows: No data")
                return None
            
            close = data['Close']
            volume = data['Volume']
            
            inflow_count = 0
            outflow_count = 0
            flow_details = []
            
            for etf in etfs:
                try:
                    if etf not in close.columns:
                        continue
                    c = close[etf].dropna()
                    v = volume[etf].dropna()
                    if len(c) < 25 or len(v) < 25:
                        continue
                    
                    # Volume surge analysis (5d avg vs 20d avg)
                    vol_5d = v.iloc[-5:].mean()
                    vol_20d = v.iloc[-20:].mean()
                    vol_change = ((vol_5d / vol_20d) - 1) * 100 if vol_20d > 0 else 0
                    
                    # Price momentum (5-day)
                    price_change = ((c.iloc[-1] / c.iloc[-5]) - 1) * 100 if c.iloc[-5] > 0 else 0
                    
                    # Flow classification
                    if vol_change > 10 and price_change > 0:
                        inflow_count += 1
                    elif vol_change > 10 and price_change < 0:
                        outflow_count += 1
                    
                    flow_details.append(f"{etf}:{vol_change:+.0f}v/{price_change:+.1f}p")
                except:
                    continue
            
            neutral_count = len(etfs) - inflow_count - outflow_count
            print(f"   ✓ ETF Flows: {inflow_count} in, {outflow_count} out, {neutral_count} neutral")
            
            return (inflow_count, outflow_count)
            
        except Exception as e:
            print(f"   ✗ ETF Flows: Error - {e}")
            return None
    
    def _credit_flow_stress(self):
        """Credit Market Flow Stress: HYG, LQD, JNK, EMB flow analysis
        Credit flows lead equity flows by 1-2 weeks.
        Uses batch download + calibrated thresholds.
        
        Classification:
        - INFLOW: volume above normal AND positive price = healthy credit demand
        - OUTFLOW: volume above normal AND negative price = credit redemption stress
        - Neutral: no strong signal
        
        Thresholds: 8% vol surge + 0.5% price move (credit ETFs have lower vol/price
        ranges than equity ETFs — old 15%/1% thresholds missed real stress signals)
        """
        try:
            credit_etfs = ['HYG', 'LQD', 'JNK', 'EMB']
            
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                data = yf.download(credit_etfs, period='2mo', progress=False, threads=True)
            
            if data.empty:
                print(f"   ✗ Credit Flows: No data")
                return None
            
            close = data['Close']
            volume = data['Volume']
            
            inflow_count = 0
            outflow_count = 0
            
            for etf in credit_etfs:
                try:
                    if etf not in close.columns:
                        continue
                    c = close[etf].dropna()
                    v = volume[etf].dropna()
                    if len(c) < 25 or len(v) < 25:
                        continue
                    
                    # Volume surge (5d vs 20d)
                    vol_5d = v.iloc[-5:].mean()
                    vol_20d = v.iloc[-20:].mean()
                    vol_change = ((vol_5d / vol_20d) - 1) * 100 if vol_20d > 0 else 0
                    
                    # Price change (5-day)
                    price_change = ((c.iloc[-1] / c.iloc[-5]) - 1) * 100 if c.iloc[-5] > 0 else 0
                    
                    # Credit-calibrated thresholds (lower than equity ETFs)
                    # Credit ETFs: HYG/LQD typically have 0.3-0.5% daily moves
                    # 0.5% over 5 days is a meaningful credit move
                    if vol_change > 8 and price_change > 0.5:
                        inflow_count += 1
                    elif vol_change > 8 and price_change < -0.5:
                        outflow_count += 1
                except:
                    continue
            
            neutral_count = len(credit_etfs) - inflow_count - outflow_count
            print(f"   ✓ Credit Flows: {inflow_count} in, {outflow_count} out, {neutral_count} neutral")
            
            return (inflow_count, outflow_count)
            
        except Exception as e:
            print(f"   ✗ Credit Flows: Error - {e}")
            return None
    
    def _sector_rotation_strength(self):
        """Sector Rotation Strength: 11-sector momentum dispersion
        Measures capital rotation intensity via top 3 vs bottom 3 momentum spread.
        
        Uses pure 10-day price momentum (no volume component).
        Rationale: during broad selloffs, volume declines across ALL sectors equally,
        making the vol component a universal drag that obscures real rotation signals.
        Price momentum alone captures where capital is flowing.
        
        Wide spread = active rotation (healthy, even in corrections)
        Narrow spread = correlated moves (panic selling or complacent buying)
        
        Uses batch yf.download() for efficiency.
        """
        try:
            sectors = ['XLK', 'XLF', 'XLV', 'XLE', 'XLY', 'XLP', 'XLI', 'XLB', 'XLRE', 'XLU', 'XLC']
            
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                data = yf.download(sectors, period='2mo', progress=False, threads=True)
            
            if data.empty:
                print(f"   ✗ Sector Rotation: No data")
                return None
            
            close = data['Close']
            sector_momentum = []
            
            for etf in sectors:
                try:
                    if etf not in close.columns:
                        continue
                    c = close[etf].dropna()
                    if len(c) < 15:
                        continue
                    
                    # 10-day price momentum (percentage)
                    momentum = ((c.iloc[-1] / c.iloc[-10]) - 1) * 100
                    sector_momentum.append(momentum)
                except:
                    continue
            
            if len(sector_momentum) >= 6:
                sector_momentum.sort(reverse=True)
                top3_avg = sum(sector_momentum[:3]) / 3
                bottom3_avg = sum(sector_momentum[-3:]) / 3
                rotation_spread = top3_avg - bottom3_avg
                
                print(f"   ✓ Sector Rotation: {rotation_spread:.1f} pts spread (top3: {top3_avg:+.1f}%, bot3: {bottom3_avg:+.1f}%)")
                return float(rotation_spread)
            else:
                print(f"   ✗ Sector Rotation: Insufficient data ({len(sector_momentum)} sectors)")
                return None
                
        except Exception as e:
            print(f"   ✗ Sector Rotation: Error - {e}")
            return None
    
    def _breadth_extreme_adjustment(self, pct_above_50ma):
        """Detect extreme overbought/oversold conditions for score adjustment
        >80% = Overbought penalty, <30% = Oversold bonus
        """
        if pct_above_50ma is None:
            return 0
        
        if pct_above_50ma > 80:
            return -5  # Overbought penalty
        elif pct_above_50ma < 30:
            return 3   # Oversold bonus (opportunity)
        else:
            return 0   # Normal range
    
    def fetch_all_data(self):
        """Fetch all 22 market indicators from FRED and yfinance APIs"""
        print("\n📊 Fetching 22 indicators...\n")
        self.missing_signals = []
        
        # Batch breadth scan — S&P 500 (replaces individual ticker loops)
        breadth = self._batch_breadth_scan()
        
        self.data = {
            'hy_spread': self._fred_get('BAMLH0A0HYM2', 'HY Spread'),
            'ig_spread': self._fred_get('BAMLC0A0CM', 'IG Spread'),
            'fed_bs_yoy': self._fed_bs_yoy(),
            'nfci': self._fred_get('NFCI', 'NFCI (Fin Conditions)'),  # Replaced TEDRATE (LIBOR deprecated)
            'dxy_trend': self._dxy_trend(),
            'pct_above_50ma': breadth.get('pct_above_50ma'),
            'pct_below_200ma': breadth.get('pct_below_200ma'),
            'ad_line': self._ad_line_status(batch_close=breadth.get('_batch_close')),
            'new_hl': breadth.get('new_hl'),
            'sector_rot': self._sector_rotation(),
            'gold_spy': self._gold_spy_ratio(),
            'yield_curve': self._fred_get('T10Y2Y', 'Yield Curve'),
            'vix': self._yf_get('^VIX', 'VIX'),
        }
        
        # Calculate credit stress ratio (HY/IG)
        if self.data.get('hy_spread') and self.data.get('ig_spread'):
            self.data['credit_stress_ratio'] = self.data['hy_spread'] / self.data['ig_spread']
        else:
            self.data['credit_stress_ratio'] = None
        
        # VIX Term Structure and SKEW
        vix_term_slope, vix_term_struct = self._vix_term_structure()
        self.data['vix_term_slope'] = vix_term_slope
        self.data['vix_term_struct'] = vix_term_struct
        self.data['skew'] = self._skew_index()
        
        # Advanced Options Intelligence
        self.data['vvix'] = self._vvix_index()
        self.data['vix9d_ratio'] = self._vix9d_ratio()
        self.data['vxn_vix_spread'] = self._vxn_vix_spread()
        
        # Institutional Flow Tracking
        self.data['etf_flows'] = self._etf_flow_divergence()
        self.data['credit_flows'] = self._credit_flow_stress()
        self.data['sector_rotation_strength'] = self._sector_rotation_strength()
        
        # Count only numeric indicators (exclude string labels like vix_term_struct and tuples)
        valid = sum(1 for k, v in self.data.items() if v is not None and k != 'vix_term_struct')
        total = len([k for k in self.data.keys() if k != 'vix_term_struct'])
        
        # Track data quality
        self.data_quality = {
            'total': total,
            'valid': valid,
            'success_rate': (valid / total * 100) if total > 0 else 0,
            'failed': self.missing_signals.copy()
        }
        
        if valid == total:
            print(f"\n✅ Fetched {valid}/{total} signals successfully (100%)\n")
        else:
            print(f"\n⚠️  Fetched {valid}/{total} signals ({self.data_quality['success_rate']:.1f}%)")
            print(f"    Missing: {', '.join(self.missing_signals)}\n")
        
        return self.data
    
    def calculate_scores(self):
        """Calculate risk scores across 3 tiers (Credit+Macro, Positioning+InstFlows, Options+Structure)
        
        SCORING CONVENTION for _score_range(val, thresholds, default):
        - Thresholds are checked left-to-right: first match (val < thresh) returns that score
        - Order thresholds LOWEST value first → HIGHEST value last
        - Assign LOW scores to LOW thresholds (bearish conditions)
        - Assign HIGH scores to HIGH thresholds (bullish conditions)
        - For "low value = good" indicators (spreads, VIX): use inverse=True
        """
        d = self.data
        
        # =====================================================================
        # TIER 1: Credit + Macro (max 39.25 pts)
        # =====================================================================
        
        # 1. HY Spread: Low spread = healthy credit = bullish (inverse: high value = bad)
        s1 = self._score_range(d.get('hy_spread'), [(3,20),(4,16),(4.5,12),(5.5,6)], 0, neutral_on_missing=True)
        hy_change = self.history_manager.get_spread_change('hy_spread', 30)
        if hy_change and hy_change['change_pct'] > 20:  # Rapid widening = stress
            s1 = max(0, s1 - 5)
        
        # 2. IG Spread: Low spread = healthy (inverse: high value = bad)
        s_ig = self._score_range(d.get('ig_spread'), [(1.2,20),(1.5,16),(2.0,12),(2.5,6)], 0, neutral_on_missing=True)
        ig_change = self.history_manager.get_spread_change('ig_spread', 30)
        if ig_change and ig_change['change_pct'] > 20:
            s_ig = max(0, s_ig - 5)
        
        # 3. Credit Stress Ratio (HY/IG): Low ratio = healthy
        # Historical: 2.5-3.5 = tight, 3.5-4.0 = normal, 4.0-4.5 = elevated, 4.5+ = stress
        s_ratio = self._score_range(d.get('credit_stress_ratio'), [(3.0,20),(3.5,16),(4.0,12),(4.5,6),(5.0,2)], 0, neutral_on_missing=True)
        
        # 4. NFCI (Chicago Fed Financial Conditions): Negative = loose (healthy), Positive = tight (stress)
        # Replaced TEDRATE (TED Spread) — LIBOR was deprecated in 2023, TED always reads near-zero
        # NFCI is composite of 105 indicators: money markets, debt, equity, banking
        # Institutional gold standard for funding conditions
        s_nfci = self._score_range(d.get('nfci'), [(-0.5,15),(-0.25,12),(0,8),(0.25,4),(0.5,2)], 0, neutral_on_missing=True)
        
        # Combine credit scores (weighted: HY 30%, IG 30%, Ratio 25%, NFCI 15%)
        credit_score = (s1 * 0.30) + (s_ig * 0.30) + (s_ratio * 0.25) + (s_nfci * 0.15)
        
        # 5. Fed Balance Sheet YoY: Positive growth = QE = bullish, negative = QT = bearish
        # Thresholds: lowest first so -12% (heavy QT) hits 4pts, >10% (QE) hits 15pts
        s2 = self._score_range(d.get('fed_bs_yoy'), [(-10,4),(-2,9),(2,12),(10,15)], 15)
        
        # 6. DXY Trend: Weak dollar = bullish for risk assets
        # Negative = dollar weakening (good), positive = strengthening (bad)
        s4 = self._score_range(d.get('dxy_trend'), [(-3,5),(-1,4),(1,3),(3,1)], 0)
        
        # =====================================================================
        # TIER 2: Positioning + Institutional Flows (max 61 pts)
        # =====================================================================
        
        # 7. % Stocks Above 50MA: High = broad rally = bullish
        # Thresholds: lowest first so 13% hits 0pts (collapse), >75% hits 12pts
        s5 = self._score_range(d.get('pct_above_50ma'), [(20,0),(35,2),(45,4),(55,7),(65,10),(75,12)], 12)
        # Remove the misleading "oversold bonus" — 13% above 50MA is NOT an opportunity,
        # it's a market in freefall. The bonus was adding +3 to a collapse reading.
        
        # 8. % Stocks Below 200MA: Low = healthy market, High = bear market
        # Use inverse=True: check val > thresh. Thresholds: highest (worst) first
        s6 = self._score_range(d.get('pct_below_200ma'), [(65,1),(50,3),(35,6),(25,8),(15,10)], 10, inverse=True)
        
        # 9. A/D Breadth: Real S&P 500 advance/decline ratio (5-day average)
        # Based on actual daily advances vs declines across 500 stocks
        # >0.55 = healthy breadth, 0.45-0.55 = mixed, <0.45 = weak/diverging
        s7 = self._score_range(d.get('ad_line'), [(0.35,0),(0.40,1),(0.45,2),(0.50,3),(0.55,4),(0.60,5)], 5)
        
        # 10. New Highs minus Lows: Positive = bullish, Negative = bearish
        # Thresholds: lowest first so -20 hits 0pts, >10 hits 3pts
        s8 = self._score_range(d.get('new_hl'), [(-10,0.5),(-5,1),(0,2),(5,2.5),(10,3)], 3)
        
        # 11. Sector Rotation: Defensive/Cyclical basket ratio trend
        # Positive = defensives outperforming (risk-off), Negative = cyclicals leading (risk-on)
        # Uses 3v3 basket (XLU+XLP+XLV vs XLY+XLF+XLI) — not a single pair
        s9 = self._score_range(d.get('sector_rot'), [(-5,6),(-2,4),(2,2),(5,0)], 0)
        
        # 12. Gold/SPY Ratio: Gold outperforming = fear/uncertainty = bearish for stocks
        # Positive = gold leading (risk-off), Negative = SPY leading (risk-on)
        s10 = self._score_range(d.get('gold_spy'), [(-5,4),(-1,3),(1,2),(5,1)], 0)
        # Regime check: during stress (VIX > 25), gold lagging SPY is likely
        # forced liquidation / margin-call selling, NOT genuine risk-on.
        # Cap bullish read at 2/4 to avoid false signal.
        vix_val = d.get('vix')
        gold_spy_val = d.get('gold_spy')
        if vix_val is not None and gold_spy_val is not None:
            if vix_val > 25 and gold_spy_val < -1:
                s10 = min(s10, 2)
        
        # 13. ETF Flow Divergence (institutional money flow)
        etf_flows = d.get('etf_flows')
        if etf_flows is not None and isinstance(etf_flows, tuple):
            inflow_count, outflow_count = etf_flows
            if outflow_count >= 7:       # Strong outflows = risk-off
                s18 = 0
            elif outflow_count >= 5:     # Moderate outflows
                s18 = 2
            elif inflow_count >= 7:      # Strong inflows = risk-on
                s18 = 8
            elif inflow_count >= 5:      # Moderate inflows
                s18 = 6
            else:                         # Mixed/neutral
                s18 = 4
        else:
            s18 = 0
        
        # 14. Credit Market Flow Stress
        credit_flows = d.get('credit_flows')
        if credit_flows is not None and isinstance(credit_flows, tuple):
            credit_in, credit_out = credit_flows
            if credit_out >= 2:          # Multiple outflows = stress
                s19 = 0
            elif credit_out >= 1:        # Some stress
                s19 = 2
            elif credit_in >= 3:         # All inflows = credit healthy
                s19 = 7
            elif credit_in >= 2:         # Mostly inflows
                s19 = 5
            else:                         # Neutral
                s19 = 3
        else:
            s19 = 0
        
        # 15. Sector Rotation Strength
        rotation_spread = d.get('sector_rotation_strength')
        if rotation_spread is not None:
            if rotation_spread > 20:     # Strong rotation = healthy bull
                s20 = 6
            elif rotation_spread > 10:   # Moderate rotation
                s20 = 4
            else:                         # Weak rotation = choppy/bear
                s20 = 0
        else:
            s20 = 0
        
        # =====================================================================
        # TIER 3: Options Intelligence + Structure (max 46 pts)
        # =====================================================================
        
        # 16. VIX Term Structure: Contango = calm, Backwardation = stress
        # Most predictive vol signal — ~35% of vol complex weight (15/46)
        vix_term = d.get('vix_term_slope')
        if vix_term is not None:
            if vix_term < -5:          # Backwardation = fear/stress
                s11 = 0
            elif vix_term < 5:         # Flat = uncertainty
                s11 = 5
            elif vix_term < 15:        # Mild contango = healthy
                s11 = 10
            else:                       # Strong contango = calm/complacent
                s11 = 15
        else:
            s11 = 0
        
        # 17. Yield Curve (10Y-2Y): Positive = normal = bullish
        # Thresholds: lowest first so deeply inverted hits low score
        s12 = self._score_range(d.get('yield_curve'), [(-0.5,0),(-0.2,1),(0.2,2),(0.5,2.5)], 3)
        
        # 18. VIX Level: Low = calm = bullish, High = fear = bearish
        # Confirmatory signal — ~22% of vol complex weight (10/46)
        s13 = self._score_range(d.get('vix'), [(15,10),(20,7),(25,4),(30,2),(35,1)], 0, neutral_on_missing=True)
        
        # 19. SKEW Index: Measures tail risk demand
        # Secondary tail risk signal — ~11% of vol complex weight (5/46)
        # Normal range 120-140 = healthy. Very low = complacency, very high = fear
        skew = d.get('skew')
        if skew is not None:
            if skew < 110:         # Extreme complacency (no hedging)
                s14 = 1
            elif skew < 120:       # Low caution
                s14 = 3
            elif skew < 140:       # Normal/healthy hedging
                s14 = 5
            elif skew < 150:       # Elevated fear
                s14 = 3
            else:                   # Extreme fear (tail risk panic)
                s14 = 1
        else:
            s14 = 0
        
        # 20. VVIX (VIX of VIX): Measures uncertainty about volatility
        # Normal 80-100. Very high = crisis uncertainty = BEARISH (not opportunity)
        vvix = d.get('vvix')
        if vvix is not None:
            if vvix < 75:          # Dangerous complacency
                s15 = 2
            elif vvix < 90:        # Calm (bullish)
                s15 = 8
            elif vvix < 100:       # Normal
                s15 = 6
            elif vvix < 115:       # Elevated uncertainty
                s15 = 3
            else:                   # Extreme uncertainty = DANGER
                s15 = 0
        else:
            s15 = 0
        
        # 21. VIX9D/VIX Ratio: Short-term fear premium
        # Event premium signal — ~7% of vol complex weight (3/46)
        vix9d_ratio = d.get('vix9d_ratio')
        if vix9d_ratio is not None:
            if vix9d_ratio > 15:       # Extreme near-term fear = stress
                s16 = 0
            elif vix9d_ratio > 5:      # Elevated near-term risk
                s16 = 1
            elif vix9d_ratio > -5:     # Normal range
                s16 = 3
            else:                       # Near-term calm (complacency or post-event)
                s16 = 2
        else:
            s16 = 0
        
        # 22. VXN-VIX Spread: Tech fear premium over market
        # Supplementary sector divergence — ~4% of vol complex weight (2/46)
        vxn_vix = d.get('vxn_vix_spread')
        if vxn_vix is not None:
            if vxn_vix > 6:            # Extreme tech fear (tech selling)
                s17 = 0
            elif vxn_vix > 3:          # Elevated tech fear
                s17 = 0.5
            elif vxn_vix > -2:         # Balanced
                s17 = 2
            else:                       # Tech complacency (tech rally)
                s17 = 1
        else:
            s17 = 0
        
        # =====================================================================
        # TOTAL SCORE
        # =====================================================================
        total = credit_score + s2 + s4 + s5 + s6 + s7 + s8 + s9 + s10 + s11 + s12 + s13 + s14 + s15 + s16 + s17 + s18 + s19 + s20
        
        # Normalize to /100 scale (max raw score is 146.25 pts)
        # T1=39.25, T2=61, T3=46 = 146.25 pts
        # T3 rebalanced to institutional vol complex hierarchy:
        #   VIX Term(15) > VIX Level(10) > VVIX(8) > SKEW(5) > Yield(3) > VIX9D(3) > VXN(2)
        SCORE_NORMALIZATION = 146.25 / 100
        total_normalized = total / SCORE_NORMALIZATION
        
        self.scores = {
            'total': total_normalized,  # Normalized to /100
            'total_raw': total,  # Keep raw for debugging
            'tier1': credit_score + s2 + s4,  # Credit composite + Fed BS + DXY
            'tier2': s5 + s6 + s7 + s8 + s9 + s10 + s18 + s19 + s20,  # Positioning + institutional flows
            'tier3': s11 + s12 + s13 + s14 + s15 + s16 + s17,  # Options intelligence + structure
            'credit_score': credit_score,
            's1': s1,  # HY Spread score
            's_ig': s_ig,  # IG Spread score
            's_ratio': s_ratio,  # Credit Stress Ratio score
            's_nfci': s_nfci,  # NFCI score
            'vix_term': s11,  # VIX Term Structure score
            'skew_score': s14,  # SKEW Index score
            'vvix_score': s15,  # VVIX score
            'vix9d_score': s16,  # VIX9D ratio score
            'vxn_score': s17,  # VXN-VIX spread score
            'etf_flow_score': s18,  # ETF flow divergence score
            'credit_flow_score': s19,  # Credit flow stress score
            'rotation_score': s20,  # Sector rotation strength score
            's2': s2, 's4': s4, 's5': s5, 's6': s6, 's7': s7, 's8': s8,
            's9': s9, 's10': s10, 's11': s11, 's12': s12, 's13': s13, 's14': s14,
            's15': s15, 's16': s16, 's17': s17, 's18': s18, 's19': s19, 's20': s20
        }
        return self.scores
    
    def _score_range(self, val, thresholds, default, inverse=False, neutral_on_missing=False):
        """Score based on thresholds. If neutral_on_missing=True, use midpoint instead of 0 for missing data."""
        if val is None:
            if neutral_on_missing and thresholds:
                # Use neutral score (average of min/max possible scores)
                scores = [s for _, s in thresholds]
                return sum(scores) / len(scores) if scores else default
            return default
        for thresh, score in thresholds:
            if (val < thresh if not inverse else val > thresh):
                return score
        return default
    
    # =============================================================================
    # IBKR POSITION INTEGRATION
    # =============================================================================
    
    def load_actual_positions(self):
        """Load actual positions from IBKR Dashboard (Column K has combined HK+AL totals)"""
        try:
            import openpyxl
            from pathlib import Path
            
            dashboard_path = Path(__file__).parent / 'fetch-ibkr-positions-dashboard.xlsx'
            
            if not dashboard_path.exists():
                return None
            
            # Read values directly from Dashboard sheet Column K (HK + AL combined)
            wb = openpyxl.load_workbook(dashboard_path, data_only=True)
            ws = wb['Dashboard']
            
            # Column K contains the combined totals (F + I columns)
            positions = {
                'global_triads': ws['K4'].value or 0,      # Global Triads
                'four_horsemen': ws['K11'].value or 0,     # Four Horsemen
                'cash_cow': ws['K19'].value or 0,          # Cash Cow
                'alpha': ws['K20'].value or 0,             # The Alpha
                'omega': ws['K21'].value or 0,             # The Omega
                'vault': ws['K22'].value or 0,             # The Vault
                'war_chest': ws['K23'].value or 0,         # The War Chest
                'total': ws['G42'].value or 0,             # Combined Total
            }
            
            wb.close()
            
            total_value = positions['total']
            if total_value == 0:
                return None
            
            print(f"   📊 Portfolio loaded from Dashboard: ${total_value:,.0f}")
            
            # Convert to percentages for compatibility with rest of code
            position_details = {
                'total_value': total_value,
                'by_category': {},
                'by_symbol': {}
            }
            
            positions['position_details'] = position_details
            
            # Convert dollar values to percentages
            for key in positions:
                if key not in ['total', 'position_details']:
                    positions[key] = positions[key] / total_value
            
            self.actual_positions = positions
            return positions
            
        except Exception as e:
            print(f"   ⚠️  Could not load IBKR positions: {e}")
            return None
    
    def compare_to_target(self, target_allocation):
        """Compare actual positions to target allocation"""
        if not self.actual_positions:
            return None
        
        actual = self.actual_positions
        drift = {}
        total_drift = 0
        alerts = []
        
        # Calculate drift for each category
        print(f"🎯 DRIFT CALCULATION:")
        for category in ['global_triads', 'four_horsemen', 'cash_cow', 'alpha', 'omega', 'vault', 'war_chest']:
            target_pct = target_allocation.get(category, 0)
            actual_pct = actual.get(category, 0)
            diff = actual_pct - target_pct
            drift[category] = diff
            total_drift += abs(diff)
            
            print(f"   {category}: Actual {actual_pct*100:.1f}% - Target {target_pct*100:.1f}% = {diff*100:+.1f}% (abs: {abs(diff)*100:.1f}%)")
            
            # Flag significant drifts (>3%)
            if abs(diff) > 0.03:
                direction = 'overweight' if diff > 0 else 'underweight'
                alerts.append(f"{category.replace('_', ' ').title()}: {direction} by {abs(diff)*100:.0f}%")
        
        print(f"   ───────────────────────────────────")
        print(f"   TOTAL DRIFT: {total_drift*100:.1f}%\n")
        
        # Save drift snapshot to history
        today = datetime.now().strftime('%Y-%m-%d')
        self.history_manager.add_drift_snapshot(today, total_drift, drift)
        self.history_manager.save_history()
        
        return {
            'drift': drift,
            'total_drift': total_drift,
            'alerts': alerts,
            'show': total_drift > 0.05  # Only show if total drift > 5%
        }
    
    def generate_rebalancing_recommendations(self, target_allocation, drift_analysis):
        """Generate specific buy/sell recommendations to reach target allocation"""
        if not self.actual_positions or not drift_analysis['show']:
            return None
        
        actual = self.actual_positions
        total_value = actual.get('position_details', {}).get('total_value', 0)
        
        if total_value == 0:
            return None
        
        recommendations = {
            'sells': [],
            'buys': [],
            'total_to_sell': 0,
            'total_to_buy': 0
        }
        
        # Identify what to sell (overweight categories)
        for category in ['global_triads', 'four_horsemen', 'cash_cow', 'alpha', 'omega', 'vault', 'war_chest']:
            drift = drift_analysis['drift'].get(category, 0)
            
            if drift > 0.03:  # Overweight by >3%
                amount = drift * total_value
                recommendations['sells'].append({
                    'category': category,
                    'current_pct': actual[category] * 100,
                    'target_pct': target_allocation[category] * 100,
                    'drift_pct': drift * 100,
                    'amount': amount
                })
                recommendations['total_to_sell'] += amount
        
        # Identify what to buy (underweight categories)
        for category in ['global_triads', 'four_horsemen', 'cash_cow', 'alpha', 'omega', 'vault', 'war_chest']:
            drift = drift_analysis['drift'].get(category, 0)
            
            if drift < -0.03:  # Underweight by >3%
                amount = abs(drift) * total_value
                recommendations['buys'].append({
                    'category': category,
                    'current_pct': actual[category] * 100,
                    'target_pct': target_allocation[category] * 100,
                    'drift_pct': drift * 100,
                    'amount': amount
                })
                recommendations['total_to_buy'] += amount
        
        return recommendations if recommendations['sells'] or recommendations['buys'] else None
    
    def analyze_options_positions(self):
        """Analyze income strategy options for expiry and urgency"""
        try:
            import pandas as pd
            from pathlib import Path
            from datetime import datetime, timedelta
            
            excel_path = Path(__file__).parent / 'fetch-ibkr-positions.xlsx'
            if not excel_path.exists():
                return None
            
            df_hk = pd.read_excel(excel_path, sheet_name='PositionsHK')
            df_al = pd.read_excel(excel_path, sheet_name='PositionsAL')
            
            analysis = {
                'csps': [],
                'covered_calls': [],
                'spreads': [],
                'stocks': [],
                'total_premium': 0,
                'days_to_nearest_expiry': None
            }
            
            today = datetime.now()
            nearest_expiry = None
            
            # Analyze all options
            for df in [df_hk, df_al]:
                opts = df[df['AssetClass'].isin(['OPT', 'FOP'])]
                for _, row in opts.iterrows():
                    symbol = row['Symbol']
                    side = row.get('Side', '')
                    put_call = row.get('Put/Call', '')
                    value = row['PositionValueUSD']
                    expiry = row.get('Expiry', '')
                    
                    # Calculate days to expiry
                    days_to_expiry = None
                    if pd.notna(expiry):
                        try:
                            expiry_date = pd.to_datetime(expiry)
                            days_to_expiry = (expiry_date - today).days
                            
                            if nearest_expiry is None or days_to_expiry < nearest_expiry:
                                nearest_expiry = days_to_expiry
                        except:
                            pass
                    
                    # Categorize
                    if side == 'Short':
                        if put_call == 'P':
                            analysis['csps'].append({
                                'symbol': symbol,
                                'value': value,
                                'days_to_expiry': days_to_expiry
                            })
                        elif put_call == 'C':
                            analysis['covered_calls'].append({
                                'symbol': symbol,
                                'value': value,
                                'days_to_expiry': days_to_expiry
                            })
                        analysis['total_premium'] += abs(value)
                
                # Get stocks in income strategy (Cash Cow)
                income_tickers = SYMBOL_MAPPING['cash_cow']
                stocks = df[(df['AssetClass'] == 'STK') & (df['Symbol'].isin(income_tickers))]
                for _, row in stocks.iterrows():
                    analysis['stocks'].append({
                        'symbol': row['Symbol'],
                        'value': row['PositionValueUSD']
                    })
            
            analysis['days_to_nearest_expiry'] = nearest_expiry
            return analysis
            
        except Exception as e:
            print(f"   ⚠️  Could not analyze options: {e}")
            return None
    
    # =============================================================================
    # v2.0: 2026 PORTFOLIO ALLOCATION LOGIC (score normalized to /100)
    # =============================================================================
    
    def _normalize_allocation(self, alloc):
        """Normalize allocation percentages to sum to exactly 100%."""
        # Extract allocation percentages
        categories = ['global_triads', 'four_horsemen', 'cash_cow', 'alpha', 'omega', 'vault', 'war_chest']
        total = sum(alloc[cat] for cat in categories)
        
        # Normalize to 100%
        if total > 0 and abs(total - 1.0) > 0.0001:  # Only normalize if off by more than 0.01%
            for cat in categories:
                alloc[cat] = alloc[cat] / total
        
        return alloc
    
    def get_portfolio_allocation(self):
        """
        Map risk score to specific portfolio adjustments
        Returns: dict with position-specific guidance
        """
        score = self.scores['total']
        base = PORTFOLIO_2026['BASE_ALLOCATION']
        
        # Score-based allocation
        if score >= 90:  # ALL CLEAR
            base_allocation = {
                'regime': '★★★★★ ALL CLEAR',
                'global_triads': base['global_triads'],
                'four_horsemen': base['four_horsemen'],
                'cash_cow': base['cash_cow'],
                'alpha': base['alpha'],
                'omega': base['omega'],
                'vault': base['vault'],
                'war_chest': base['war_chest'],
                'stops': '15-20%',
                'options_guidance': 'Sell CSPs at 30-45 DTE, 15-20 delta on GOOGL, PEP, V',
                'action': 'FULL DEPLOYMENT - Run base allocation'
            }
            base_allocation = self._normalize_allocation(base_allocation)
        
        elif score >= 75:  # NORMAL
            base_allocation = {
                'regime': '★★★★☆ NORMAL',
                'global_triads': base['global_triads'],
                'four_horsemen': base['four_horsemen'] * 0.9,  # Slightly reduce growth
                'cash_cow': base['cash_cow'] * 0.9,  # More conservative
                'alpha': base['alpha'] * 0.8,  # Reduce offensive plays
                'omega': 0.01,  # 1% light hedging (minimal cost)
                'vault': base['vault'] + 0.02,  # 7% gold
                'war_chest': base['war_chest'] + 0.05,  # 10% cash
                'stops': '12-15%',
                'options_guidance': 'Tighter strikes: 10-15 delta CSPs, 30 DTE. Light put hedging: 1-2 contracts only',
                'action': 'STAY COURSE - Tighten stops, minimal hedging (1%)'
            }
            base_allocation = self._normalize_allocation(base_allocation)
        
        elif score >= 60:  # ELEVATED
            base_allocation = {
                'regime': '★★★☆☆ ELEVATED',
                'global_triads': base['global_triads'],
                'four_horsemen': base['four_horsemen'] * 0.6,  # Cut growth more aggressively
                'cash_cow': base['cash_cow'] * 0.4,  # Very defensive options
                'alpha': base['alpha'] * 0.3,  # Cut offensive plays harder
                'omega': 0.01,  # 1% only (annual cost ~$3k vs $10k at 5%)
                'vault': 0.10,  # 10% gold
                'war_chest': 0.24,  # 24% cash (raise cash instead of buying puts)
                'stops': '10-12%',
                'options_guidance': 'DEFENSIVE: Far OTM CSPs (5-10 delta), close losing positions. Minimal hedging: 2-3 put spreads max',
                'action': 'REDUCE RISK - Cut positions, raise cash to 24%. Light hedging (1% only)'
            }
            base_allocation = self._normalize_allocation(base_allocation)
        
        elif score >= 40:  # HIGH RISK
            base_allocation = {
                'regime': '★★☆☆☆ HIGH RISK',
                'global_triads': base['global_triads'] * 0.7,  # Trim core more
                'four_horsemen': base['four_horsemen'] * 0.2,  # Skeleton growth
                'cash_cow': 0,  # Exit all income strategies
                'alpha': 0,  # Exit all offensive plays
                'omega': 0.02,  # 2% hedging (annual cost ~$6k vs $15k at 5%)
                'vault': 0.15,  # 15% gold
                'war_chest': 0.33,  # 33% cash (sell positions, don't buy puts)
                'stops': '8-10%',
                'options_guidance': 'CLOSE POSITIONS: Roll losing CSPs, collect premium and exit. Use collars (sell calls to fund puts) for cost-neutral hedging',
                'action': 'GO DEFENSIVE - Sell positions aggressively, raise cash to 33%. Moderate hedging (2%)'
            }
            base_allocation = self._normalize_allocation(base_allocation)
        
        else:  # EXTREME (<40)
            base_allocation = {
                'regime': '★☆☆☆☆ EXTREME RISK',
                'global_triads': base['global_triads'] * 0.3,  # Cut core drastically
                'four_horsemen': 0,  # Exit all growth
                'cash_cow': 0,  # No options
                'alpha': 0,  # Exit all offensive plays
                'omega': 0.03,  # 3% MAX (annual cost ~$10k - at portfolio limit)
                'vault': 0.25,  # 25% gold
                'war_chest': 0.39,  # 39% cash (maximum liquidity)
                'stops': '5-8%',
                'options_guidance': 'CLOSE ALL: Exit CSPs at any reasonable price. ONLY use collars (sell calls to fund puts) - zero net cost. VIX >40 makes puts too expensive',
                'action': 'MAX DEFENSE - Liquidate to 39% cash, 25% gold. Hedging at 3% max annual cost limit'
            }
            base_allocation = self._normalize_allocation(base_allocation)
        
        # Apply V-Recovery override if active
        if self.v_recovery_active and base_allocation:
            return self._apply_v_recovery_to_portfolio(base_allocation)
        return base_allocation if base_allocation else self.get_portfolio_allocation()  # Fallback
    
    def _apply_v_recovery_to_portfolio(self, base_alloc):
        """Apply V-Recovery override - cut reserves by 50% and redeploy"""
        old_gold = base_alloc['vault']
        old_cash = base_alloc['war_chest']
        total_reserves = old_gold + old_cash
        new_reserves = total_reserves * 0.5
        freed_capital = total_reserves - new_reserves
        
        # Redistribute freed capital proportionally to core/growth/income
        total_active = base_alloc['global_triads'] + base_alloc['four_horsemen'] + base_alloc['cash_cow']
        
        if total_active > 0:
            boost_factor = 1 + (freed_capital / total_active)
            v_alloc = {
                **base_alloc,
                'regime': base_alloc['regime'] + ' + V-RECOVERY',
                'global_triads': base_alloc['global_triads'] * boost_factor,
                'four_horsemen': base_alloc['four_horsemen'] * boost_factor,
                'cash_cow': base_alloc['cash_cow'] * boost_factor,
                'vault': old_gold * 0.5,  # Cut gold 50%
                'war_chest': old_cash * 0.5,  # Cut cash 50%
                'action': base_alloc['action'] + ' | V-RECOVERY: Reserves cut 50%, redeployed to core/growth/income'
            }
            return self._normalize_allocation(v_alloc)
        else:
            return base_alloc
    
    # =============================================================================
    # V-RECOVERY DETECTION with KILL-SWITCH
    # =============================================================================
    
    def check_v_recovery_trigger(self):
        """Check if V-Recovery override should activate based on SPY decline and VIX spike"""
        if not CONFIG['V_RECOVERY_ENABLED']:
            return False, None
        
        try:
            spy = yf.Ticker('SPY').history(period='2mo')
            if len(spy) < CONFIG['V_RECOVERY_SPY_DAYS']:
                return False, None
            
            # Check if override already active and apply kill-switch
            # Use consecutive streak logic (not calendar days) to match v1.7+ behavior
            override_streak = self.history_manager.get_override_streak()
            if override_streak >= 5:
                current_score = self.scores['total']
                if current_score < 60:
                    print(f"⚠️ KILL-SWITCH: Override active {override_streak} consecutive days but score={current_score:.1f} < 60. ABORTING.")
                    return False, f"KILL-SWITCH activated: {override_streak} days active, score={current_score:.1f}"
            
            had_extreme = self.history_manager.had_extreme_score_recently(
                threshold=CONFIG['V_RECOVERY_SCORE_THRESHOLD'],
                days=CONFIG['V_RECOVERY_LOOKBACK']
            )
            
            if not had_extreme:
                return False, None
            
            lookback = CONFIG['V_RECOVERY_SPY_DAYS']
            spy_gain = ((spy['Close'].iloc[-1] - spy['Close'].iloc[-lookback]) / 
                       spy['Close'].iloc[-lookback] * 100)
            
            if spy_gain < CONFIG['V_RECOVERY_SPY_GAIN']:
                return False, None
            
            recent_scores = self.history_manager.get_recent_scores(CONFIG['V_RECOVERY_LOOKBACK'])
            if not recent_scores or self.data.get('vix') is None:
                return False, None
            
            vix_high = max(s['vix'] for s in recent_scores if s.get('vix'))
            vix_drop = vix_high - self.data['vix']
            
            if vix_drop < CONFIG['V_RECOVERY_VIX_DROP']:
                return False, None
            
            if self.data.get('hy_spread') is not None:
                credit_stable = self.data['hy_spread'] < 6.0
                if not credit_stable:
                    return False, None
            
            reason = (
                f"V-Recovery Triggered:\n"
                f"  • Score <{CONFIG['V_RECOVERY_SCORE_THRESHOLD']} recently\n"
                f"  • SPY +{spy_gain:.1f}% in {lookback} days\n"
                f"  • VIX dropped {vix_drop:.1f} pts\n"
                f"  • Credit stable ({self.data['hy_spread']:.2f}%)"
            )
            
            return True, reason
            
        except Exception as e:
            print(f"   ⚠️  V-Recovery check failed: {e}")
            return False, None
    
    # =============================================================================
    # DIVERGENCE DETECTION
    # =============================================================================
    
    def detect_divergences(self):
        """Detect market divergences (low VIX with high credit stress or weak breadth)"""
        self.alerts = []
        d = self.data
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Helper to add alert with deduplication
        def add_alert_if_new(alert_type, severity, icon, msg, action, suppress_days=1):
            """Add alert only if not sent in last N days (prevent spam on glitches)"""
            if not self.history_manager.should_suppress_alert(alert_type, days=suppress_days):
                self.alerts.append({
                    'type': alert_type,
                    'severity': severity,
                    'icon': icon,
                    'msg': msg,
                    'action': action
                })
                self.history_manager.add_alert(today, alert_type, msg)
        
        if d.get('vix') and d['vix'] < 15:
            if (d.get('hy_spread') and d['hy_spread'] > 4.5) or \
               (d.get('pct_above_50ma') and d['pct_above_50ma'] < 50):
                add_alert_if_new(
                    'HIDDEN DANGER',
                    'CRITICAL',
                    '🚨🚨🚨',
                    'VIX CALM BUT CREDIT/BREADTH DETERIORATING',
                    'REDUCE RISK NOW',
                    suppress_days=1  # Alert once per day max
                )
        
        if (d.get('fed_bs_yoy') and d['fed_bs_yoy'] < -5) and \
           (d.get('dxy_trend') and d['dxy_trend'] > 3):
            add_alert_if_new(
                'LIQUIDITY DRAIN',
                'HIGH',
                '🚨🚨',
                'FED CONTRACTING + DOLLAR SURGING',
                'REDUCE RISK ASSETS 20-30%',
                suppress_days=2  # Alert every 2 days if condition persists
            )
        
        if (d.get('hy_spread') and d['hy_spread'] > 5) or \
           (d.get('nfci') and d['nfci'] > 0.5):
            add_alert_if_new(
                'CREDIT WARNING',
                'HIGH',
                '🚨🚨',
                'CREDIT MARKETS PRICING STRESS',
                'GO DEFENSIVE',
                suppress_days=1
            )
        
        if d.get('vix_struct') == 'Backwardation':
            streak, avg_mag = self.history_manager.get_backwardation_streak()
            if streak >= 5:
                add_alert_if_new(
                    'BACKWARDATION PERSISTING',
                    'CRITICAL',
                    '🚨🚨🚨',
                    f'VIX BACKWARDATION DAY {streak}',
                    'TIGHTEN STOPS, REDUCE TIER 3',
                    suppress_days=0  # Alert every day for critical condition
                )
            elif streak >= 3:
                add_alert_if_new(
                    'BACKWARDATION PATTERN',
                    'HIGH',
                    '🚨🚨',
                    f'VIX BACKWARDATION DAY {streak}',
                    'WATCH CREDIT & BREADTH',
                    suppress_days=1
                )
            elif streak >= 1 and d.get('vix') and d['vix'] < 16:
                add_alert_if_new(
                    'HIDDEN TENSION',
                    'MEDIUM',
                    '⚠️',
                    f'VIX CALM BUT BACKWARDATION DETECTED',
                    'INSTITUTIONS BUYING PROTECTION',
                    suppress_days=2
                )
        
        if self.v_recovery_active:
            add_alert_if_new(
                'V-RECOVERY ACTIVE',
                'INFO',
                '🚀',
                self.v_recovery_reason,
                'AGGRESSIVE RE-ENTRY',
                suppress_days=1
            )
        
        if not self.alerts and self.scores['total'] >= 85:
            # Only send ALL CLEAR once every 7 days (don't spam)
            if not self.history_manager.should_suppress_alert('ALL CLEAR', days=7):
                self.alerts.append({
                    'type': 'ALL CLEAR',
                    'severity': 'SAFE',
                    'icon': '✅',
                    'msg': 'HEALTHY MARKET CONDITIONS',
                    'action': 'FULL DEPLOYMENT OK'
                })
                self.history_manager.add_alert(today, 'ALL CLEAR', 'HEALTHY MARKET CONDITIONS')
        
        return self.alerts
    
    # =============================================================================
    # REPORTING with 2026 PORTFOLIO BREAKDOWN
    # =============================================================================
    
    def generate_report(self):
        """Generate comprehensive risk report with scores, allocations, and recommendations"""
        score = self.scores['total']
        d = self.data
        
        # Get portfolio allocation
        portfolio = self.get_portfolio_allocation()
        
        # Check V-Recovery
        self.v_recovery_active, self.v_recovery_reason = self.check_v_recovery_trigger()
        if self.v_recovery_active:
            portfolio = self.get_portfolio_allocation()  # Recalc with override
        
        lines = [
            "🎯 RISK DASHBOARD v2.0",
            f"📅 {self.timestamp.strftime('%b %d, %Y @ %H:%M')}",
            "",
        ]
        
        # Data Quality Report
        dq = self.data_quality
        if dq['success_rate'] == 100:
            lines.append(f"✅ DATA: {dq['valid']}/{dq['total']} signals (100%)")
        else:
            lines.append(f"⚠️  DATA: {dq['valid']}/{dq['total']} signals ({dq['success_rate']:.0f}%)")
            if dq['failed']:
                lines.append(f"   Missing: {', '.join(dq['failed'][:3])}{'...' if len(dq['failed']) > 3 else ''}")
        lines.append("")
        
        lines.extend([
            f"📊 SCORE: {score:.1f}/100",
            f"🎚️ {portfolio['regime']}",
            "",
        ])
        
        # Load actual positions and compare to target
        self.actual_positions = self.load_actual_positions()
        drift_analysis = self.compare_to_target(portfolio) if self.actual_positions else None
        
        # Get rebalancing recommendations and options analysis
        rebalance_recs = self.generate_rebalancing_recommendations(portfolio, drift_analysis) if drift_analysis else None
        options_analysis = self.analyze_options_positions() if self.actual_positions else None
        
        # Get drift trend and score trend
        drift_trend = self.history_manager.get_drift_trend(days=7) if drift_analysis else None
        score_trend = self.history_manager.get_score_trend(days=7)
        
        # Portfolio allocation breakdown
        if drift_analysis and drift_analysis['show']:
            # Show actual vs target with action needed
            lines.extend(["💼 PORTFOLIO ALLOCATION (Actual → Target)"])
            
            categories = [
                ('Global Triads', 'global_triads'),
                ('Four Horsemen', 'four_horsemen'),
                ('Cash Cow', 'cash_cow'),
                ('The Alpha', 'alpha'),
                ('The Omega', 'omega'),
                ('The Vault', 'vault'),
                ('War Chest', 'war_chest')
            ]
            
            for label, key in categories:
                target = portfolio[key] * 100
                actual = self.actual_positions[key] * 100
                diff = drift_analysis['drift'][key] * 100
                
                if abs(diff) > 3:
                    indicator = "⚠️" if abs(diff) > 5 else "⚡"
                    # Show action needed: negative diff means underweight (need to add +), positive means overweight (need to reduce -)
                    action = f"{-diff:+.1f}%"
                    lines.append(f"{label}: {actual:.1f}% → {target:.1f}% ({action} {indicator})")
                else:
                    lines.append(f"{label}: {actual:.1f}% → {target:.1f}%")
            
            # Show drift with actionable guidance
            current_drift = drift_analysis['total_drift'] * 100
            drift_line = f"⚠️ DRIFT: {current_drift:.0f}% total"
            
            if drift_trend:
                days = drift_trend['days']
                velocity = abs(drift_trend['change'] * 100) / days if days > 0 else 0
                
                if drift_trend['improving']:
                    if current_drift < 10:
                        # Almost aligned - celebrate progress
                        drift_line = f"⚠️ DRIFT: {current_drift:.0f}% total ✅ Nearly aligned (target: <5%)"
                    elif velocity > 5:
                        # Improving fast - encourage continuation
                        drift_line = f"⚠️ DRIFT: {current_drift:.0f}% total 📉 Rebalancing working - keep deploying cash (was {drift_analysis['total_drift']*100 + drift_trend['change']*100:.0f}%)"
                    else:
                        # Improving slowly - maintain course
                        drift_line = f"⚠️ DRIFT: {current_drift:.0f}% total ↘️ Improving but still high (target: <5%)"
                else:
                    if velocity > 3:
                        # Worsening fast - urgent action
                        drift_line = f"⚠️ DRIFT: {current_drift:.0f}% total 🚨 URGENT: Increasing fast (+{drift_trend['change']*100:.0f}% in {days}d)"
                    else:
                        # Worsening slowly - warning
                        drift_line = f"⚠️ DRIFT: {current_drift:.0f}% total ⚠️ Worsening - review positions"
            
            lines.extend([
                "",
                drift_line,
                f"📋 {', '.join(drift_analysis['alerts'][:2])}",  # Show top 2 alerts
                ""
            ])
        else:
            # Show target allocation only
            lines.extend([
                "💼 PORTFOLIO ALLOCATION (Target)",
                f"Global Triads: {portfolio['global_triads']*100:.0f}%",
                f"Four Horsemen: {portfolio['four_horsemen']*100:.0f}%",
                f"Cash Cow: {portfolio['cash_cow']*100:.0f}%",
                f"The Alpha: {portfolio['alpha']*100:.0f}%",
                f"The Omega: {portfolio['omega']*100:.0f}%",
                f"The Vault: {portfolio['vault']*100:.0f}%",
                f"War Chest: {portfolio['war_chest']*100:.0f}%",
                ""
            ])
        
        lines.extend([
            f"🛡️ Stops: {portfolio['stops']}",
            f"📈 Options: {portfolio['options_guidance']}",
            "",
            f"💡 ACTION: {portfolio['action']}",
            "",
        ])
        
        # Add rebalancing recommendations if significant drift
        if rebalance_recs and (rebalance_recs['sells'] or rebalance_recs['buys']):
            lines.extend(["🔄 REBALANCING RECOMMENDATIONS"])
            
            if rebalance_recs['sells']:
                lines.append("SELL (Overweight):")
                for rec in rebalance_recs['sells'][:3]:  # Top 3
                    cat_name = rec['category'].replace('_', ' ').title()
                    lines.append(f"  • {cat_name}: ${rec['amount']/1000:.0f}k ({rec['drift_pct']:.0f}% over)")
            
            if rebalance_recs['buys']:
                lines.append("BUY (Underweight):")
                for rec in rebalance_recs['buys'][:3]:  # Top 3
                    cat_name = rec['category'].replace('_', ' ').title()
                    lines.append(f"  • {cat_name}: ${rec['amount']/1000:.0f}k ({abs(rec['drift_pct']):.0f}% under)")
            
            lines.append("")
        
        # Add options analysis if available
        if options_analysis:
            total_csps = len(options_analysis['csps'])
            total_ccs = len(options_analysis['covered_calls'])
            total_stocks = len(options_analysis['stocks'])
            days_to_expiry = options_analysis['days_to_nearest_expiry']
            
            if total_csps > 0 or total_ccs > 0 or total_stocks > 0:
                lines.extend(["📊 INCOME STRATEGY POSITIONS"])
                
                if total_csps > 0:
                    lines.append(f"  • CSPs: {total_csps} active")
                if total_ccs > 0:
                    lines.append(f"  • Covered Calls: {total_ccs} active")
                if total_stocks > 0:
                    stock_value = sum(s['value'] for s in options_analysis['stocks'])
                    lines.append(f"  • Stock Holdings: {total_stocks} positions (${stock_value/1000:.0f}k)")
                
                if days_to_expiry is not None:
                    urgency = "🔴 URGENT" if days_to_expiry <= 3 else "🟡 Soon" if days_to_expiry <= 7 else "🟢 OK"
                    lines.append(f"  • Nearest Expiry: {days_to_expiry} days {urgency}")
                
                lines.append("")
        
        # Tier scores with underlying indicator breakdown
        s = self.scores
        lines.extend([
            "📈 TIER SCORES",
            f"T1: {s['tier1']:.1f}/39.25 Credit+Macro ({s['tier1']/39.25*100:.0f}%)",
            f"  Credit {s['credit_score']:.1f}/19.25: HY {s['s1']:.0f}/20 · IG {s['s_ig']:.0f}/20 · Ratio {s['s_ratio']:.0f}/20 · NFCI {s['s_nfci']:.0f}/15",
            f"  Fed BS {s['s2']:.0f}/15 · DXY {s['s4']:.0f}/5",
            f"T2: {s['tier2']:.1f}/61 Positioning+Flows ({s['tier2']/61*100:.0f}%)",
            f"  >50MA {s['s5']:.0f}/12 · <200MA {s['s6']:.0f}/10 · A/D {s['s7']:.0f}/5 · H-L {s['s8']:.1f}/3",
            f"  Rotation {s['s9']:.0f}/6 · Au/SPY {s['s10']:.0f}/4 · ETF {s['s18']:.0f}/8 · Credit {s['s19']:.0f}/7 · Sect {s['s20']:.0f}/6",
            f"T3: {s['tier3']:.1f}/46 Options+Structure ({s['tier3']/46*100:.0f}%)",
            f"  VIXTerm {s['s11']:.0f}/15 · YC {s['s12']:.1f}/3 · VIX {s['s13']:.0f}/10 · SKEW {s['s14']:.0f}/5",
            f"  VVIX {s['s15']:.0f}/8 · 9D/VIX {s['s16']:.0f}/3 · VXN {s['s17']:.0f}/2",
            "",
        ])
        
        # Score trend (historical comparison)
        if score_trend and score_trend['days'] >= 2:
            lines.extend(["📊 RISK TREND ({} days)".format(score_trend['days'])])
            
            # Show score movement
            if score_trend['improving']:
                trend_icon = "📈" if score_trend['change'] > 5 else "↗️"
                lines.append(f"{trend_icon} Score: {score_trend['first_score']:.0f} → {score_trend['last_score']:.0f} (+{score_trend['change']:.0f} pts)")
            else:
                trend_icon = "📉" if score_trend['change'] < -5 else "↘️"
                lines.append(f"{trend_icon} Score: {score_trend['first_score']:.0f} → {score_trend['last_score']:.0f} ({score_trend['change']:.0f} pts)")
            
            # Show regime change if occurred
            if score_trend['regime_changed']:
                lines.append(f"⚠️ Regime: {score_trend['first_regime']} → {score_trend['last_regime']}")
            
            # Show key signal changes
            if 'vix' in score_trend['signal_changes']:
                vix = score_trend['signal_changes']['vix']
                if abs(vix['pct']) > 10:  # Only show if >10% change
                    direction = "↑" if vix['change'] > 0 else "↓"
                    lines.append(f"  • VIX: {vix['first']:.1f} → {vix['last']:.1f} ({direction}{abs(vix['pct']):.0f}%)")
            
            lines.append("")
        
        # Key signals
        lines.extend([
            "📊 KEY SIGNALS",
            f"HY Spread: {d.get('hy_spread', 'N/A'):.2f}%" if d.get('hy_spread') else "HY Spread: N/A",
            f"VIX: {d.get('vix', 'N/A'):.1f}" if d.get('vix') else "VIX: N/A",
        ])
        
        if d.get('vix_struct'):
            vix_line = f"VIX: {d['vix_struct']}"
            if d['vix_struct'] == 'Backwardation':
                streak, _ = self.history_manager.get_backwardation_streak()
                if streak > 0:
                    vix_line = f"⚠️ VIX: Back Day {streak}"
            lines.append(vix_line)
        
        lines.extend([
            f"% >50MA: {d.get('pct_above_50ma', 'N/A'):.0f}%" if d.get('pct_above_50ma') else "% >50MA: N/A",
            f"F&G: {d.get('fear_greed', 'N/A'):.0f}" if d.get('fear_greed') else "F&G: N/A",
            "",
        ])
        
        # Alerts
        if self.alerts:
            lines.append("🚨 ALERTS")
            for alert in self.alerts:
                lines.extend(["", f"{alert['icon']} {alert['type']}", f"{alert['msg']}", f"→ {alert['action']}"])
        
        return "\n".join(lines)
    
    # =============================================================================
    # AI CIO INTERPRETATIONS - Dual Analysis (Claude + Gemini)
    # =============================================================================
    
    def generate_claude_interpretation(self):
        """Claude CIO Analysis"""
        key = os.getenv('ANTHROPIC_API_KEY')
        if not key or 'YOUR_' in key:
            return None
        
        try:
            portfolio = self.get_portfolio_allocation()
            prompt = f"""You are the CIO analyzing today's risk dashboard for a $1M portfolio trader targeting 15% annual returns.

TODAY'S DATA:
Score: {self.scores['total']:.1f}/100
Regime: {portfolio['regime']}
Portfolio Allocation:
- Global Triads (Strategic Core - 82846/DHL/ES3/VWRA/VT/XMNE): {portfolio['global_triads']*100:.0f}%
- Four Horsemen (Growth Engine - CSNDX/CTEC/HEAL/INRA/GRDU): {portfolio['four_horsemen']*100:.0f}%
- Cash Cow (Income Strategy - all options EXCEPT SPY/QQQ): {portfolio['cash_cow']*100:.0f}%
- The Alpha (Speculation - theme stocks + long calls): {portfolio['alpha']*100:.0f}%
- The Omega (Insurance - SPY/QQQ bear spreads ONLY, capped 1-3% by regime): {portfolio['omega']*100:.0f}%
- The Vault (Gold - tail risk protection): {portfolio['vault']*100:.0f}%
- War Chest (Cash - primary defense mechanism): {portfolio['war_chest']*100:.0f}%
Action: {portfolio['action']}

HEDGING PHILOSOPHY: Omega capped at 1% (ELEVATED), 2% (HIGH RISK), 3% (EXTREME) to keep annual insurance cost under 3% of portfolio. Primary defense is selling positions and raising cash, not buying expensive puts.

Key Indicators:
- HY Spread: {self.data.get('hy_spread', 'N/A')}%
- VIX: {self.data.get('vix', 'N/A')}
- % Above 50MA: {self.data.get('pct_above_50ma', 'N/A')}%
- Breadth: {self.data.get('ad_line', 'N/A')}

Write brief CIO interpretation for mobile:
💭 HEADLINE (one line)
📊 SCORE QUALITY (tiers assessment)
👁️ WHAT I SEE (2-3 bullets)
🎯 REGIME (2 sentences)
💡 MY CALL (specific to positions)
🔄 FLIP TRIGGERS (specific)
⚡ BOTTOM LINE (one sentence)

<1200 chars, use numbers, be direct."""

            resp = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                json={"model": "claude-sonnet-4-20250514", "max_tokens": 2000, "messages": [{"role": "user", "content": prompt}]},
                timeout=30
            )
            if resp.status_code == 200:
                return f"🧠 CLAUDE CIO ANALYSIS\n📅 {self.timestamp.strftime('%b %d, %Y')}\n\n{resp.json()['content'][0]['text']}"
            return None
        except Exception as e:
            print(f"⚠️ Claude CIO Error: {e}")
            return None
    
    def generate_gemini_interpretation(self):
        """Gemini CIO Analysis - Second Opinion"""
        key = os.getenv('GEMINI_API_KEY')
        if not key or 'YOUR_' in key:
            return None
        
        try:
            portfolio = self.get_portfolio_allocation()
            prompt = f"""You are the Chief Investment Officer providing a second opinion on today's risk dashboard for a $1M portfolio targeting 15% annual returns.

TODAY'S MARKET SNAPSHOT:
Risk Score: {self.scores['total']:.1f}/100
Market Regime: {portfolio['regime']}

Current Portfolio Allocation:
- Global Triads (Strategic Core - diversified ETFs): {portfolio['global_triads']*100:.0f}%
- Four Horsemen (Growth Engine - thematic ETFs): {portfolio['four_horsemen']*100:.0f}%
- Cash Cow (Income Strategy - multi-leg options on mega-caps, excludes SPY/QQQ): {portfolio['cash_cow']*100:.0f}%
- The Alpha (Speculation - theme stocks, long calls): {portfolio['alpha']*100:.0f}%
- The Omega (Insurance - SPY/QQQ bear spreads, strictly capped by regime): {portfolio['omega']*100:.0f}%
- The Vault (Gold - store of value, tail risk): {portfolio['vault']*100:.0f}%
- War Chest (Cash - dry powder, primary defense): {portfolio['war_chest']*100:.0f}%

Recommended Action: {portfolio['action']}

RISK MANAGEMENT APPROACH: Institutional practice - hedge via position sizing (selling) not expensive options. Omega capped at 1-3% max depending on regime to keep annual insurance cost under 3% of portfolio. Cash raising is primary defense.

Key Market Indicators:
- High Yield Spread: {self.data.get('hy_spread', 'N/A')}% (credit stress)
- VIX: {self.data.get('vix', 'N/A')} (fear gauge)
- Stocks Above 50MA: {self.data.get('pct_above_50ma', 'N/A')}% (breadth)
- Market Breadth: {self.data.get('ad_line', 'N/A')}

Provide your CIO perspective (max 1000 chars) covering:
💭 HEADLINE: One-line market view
📊 SCORE INSIGHT: Quality assessment
👁️ KEY OBSERVATIONS: 2-3 critical points
🎯 REGIME CONTEXT: Market environment assessment
💡 PORTFOLIO STANCE: Specific position guidance
🔄 WATCH TRIGGERS: Conditions that would change your view
⚡ EXECUTIVE SUMMARY: Bottom line call

Be direct, use specific numbers, focus on actionable insights."""

            resp = requests.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={key}",
                headers={"Content-Type": "application/json"},
                json={"contents": [{"parts": [{"text": prompt}]}]},
                timeout=30
            )
            if resp.status_code == 200:
                text = resp.json()['candidates'][0]['content']['parts'][0]['text']
                return f"💎 GEMINI CIO ANALYSIS\n📅 {self.timestamp.strftime('%b %d, %Y')}\n\n{text}"
            return None
        except Exception as e:
            print(f"⚠️ Gemini CIO Error: {e}")
            return None
    
    # =============================================================================
    # MAIN EXECUTION
    # =============================================================================
    
    def run_assessment(self):
        """Execute complete daily risk assessment workflow: fetch data, calculate scores, generate reports, send to Telegram"""
        print("\n" + "="*80)
        print("STARTING DAILY RISK ASSESSMENT v2.0 - 2026 PORTFOLIO")
        print("="*80)
        
        self.fetch_all_data()
        
        # Note: We continue even with missing signals (using neutral scores)
        # This prevents complete assessment failure due to temporary API issues
        
        # Store indicator data for historical comparison
        self.history_manager.add_indicator_data(
            date=self.timestamp.strftime('%Y-%m-%d'),
            indicators={
                'hy_spread': self.data.get('hy_spread'),
                'ig_spread': self.data.get('ig_spread'),
                'credit_stress_ratio': self.data.get('credit_stress_ratio'),
                'nfci': self.data.get('nfci'),
            }
        )
        
        self.calculate_scores()
        
        # Save history
        try:
            spy_price = yf.Ticker('SPY').history(period='1d')['Close'].iloc[-1]
        except:
            spy_price = None
        
        self.history_manager.add_score(
            date=self.timestamp.strftime('%Y-%m-%d'),
            score=self.scores['total'],
            spy_price=spy_price,
            vix=self.data.get('vix'),
            vix_structure=self.data.get('vix_struct'),
            vixy_vxx_ratio=self.data.get('vixy_vxx_ratio')
        )
        
        # V-Recovery check
        self.v_recovery_active, self.v_recovery_reason = self.check_v_recovery_trigger()
        if self.v_recovery_active:
            print(f"\n🚀 {self.v_recovery_reason}")
            self.history_manager.add_override_event(
                date=self.timestamp.strftime('%Y-%m-%d'),
                reason="V-Recovery",
                conditions=self.v_recovery_reason
            )
        
        self.detect_divergences()
        report = self.generate_report()
        
        print("\n" + report + "\n")
        
        # Save and send
        with open(f"risk_report_{self.timestamp.strftime('%Y%m%d')}.txt", 'w') as f:
            f.write(report)
        self.history_manager.save_history()
        send_to_telegram(report)
        
        # Dual CIO interpretations (4-eye principle)
        claude_cio = self.generate_claude_interpretation()
        if claude_cio:
            print("\n" + claude_cio + "\n")
            send_to_telegram(claude_cio)
        
        gemini_cio = self.generate_gemini_interpretation()
        if gemini_cio:
            print("\n" + gemini_cio + "\n")
            send_to_telegram(gemini_cio)
        
        return self.scores['total']

def main():
    """Main entry point - initialize dashboard and run daily assessment"""
    print("""
╔══════════════════════════════════════════════════════════════════════╗
║              INSTITUTIONAL RISK DASHBOARD v2.0                       ║
║    22 Indicators + Institutional Flows + Options Intelligence        ║
╚══════════════════════════════════════════════════════════════════════╝
    """)
    dashboard = RiskDashboard()
    dashboard.run_assessment()
    print("\n✅ Assessment complete.\n")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ FATAL ERROR: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
