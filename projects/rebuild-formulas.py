from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

wb = load_workbook('fetch-ibkr-positions-dashboard.xlsx')
ws = wb['Dashboard']

# Define correct formulas without quotes or paths
formulas = {
    'E4': '=SUMPRODUCT(SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!F:F,GlobalETF!A:A))',
    'H4': '=SUMPRODUCT(SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!F:F,GlobalETF!A:A))',
    
    'E6': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!F:F,"CSNDX")',
    'H6': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!F:F,"CSNDX")',
    
    'E7': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!F:F,"CTEC")',
    'H7': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!F:F,"CTEC")',
    
    'E8': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!F:F,"HEAL")',
    'H8': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!F:F,"HEAL")',
    
    'E9': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!F:F,"LOCK")',
    'H9': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!F:F,"LOCK")',
    
    'E10': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!F:F,"INRA")',
    'H10': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!F:F,"INRA")',
    
    'E11': '=SUMPRODUCT(SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!F:F,IncomeStrategy!A:A))',
    'H11': '=SUMPRODUCT(SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!F:F,IncomeStrategy!A:A))',
    
    'E13': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!D:D,"STK")-SUMPRODUCT(SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!F:F,GlobalETF!A:A))-SUMPRODUCT(SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!F:F,GrowthEngine!A:A))-SUMPRODUCT(SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!F:F,IncomeStrategy!A:A))-SUMPRODUCT(SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!F:F,Gold!A:A))',
    'H13': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!D:D,"STK")-SUMPRODUCT(SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!F:F,GlobalETF!A:A))-SUMPRODUCT(SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!F:F,GrowthEngine!A:A))-SUMPRODUCT(SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!F:F,IncomeStrategy!A:A))-SUMPRODUCT(SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!F:F,Gold!A:A))',
    
    'E14': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!D:D,"OPT",[fetch-ibkr-positions.xlsx]PositionsHK!U:U,">0")+SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!D:D,"FOP",[fetch-ibkr-positions.xlsx]PositionsHK!U:U,">0")',
    'H14': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!D:D,"OPT",[fetch-ibkr-positions.xlsx]PositionsAL!U:U,">0")+SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!D:D,"FOP",[fetch-ibkr-positions.xlsx]PositionsAL!U:U,">0")',
    
    'E15': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsHK!U:U,[fetch-ibkr-positions.xlsx]PositionsHK!F:F,"GSD")',
    'H15': '=SUMIFS([fetch-ibkr-positions.xlsx]PositionsAL!U:U,[fetch-ibkr-positions.xlsx]PositionsAL!F:F,"GSD")',
}

print('Rebuilding formulas with correct relative reference syntax...\n')

for cell_ref, formula in formulas.items():
    ws[cell_ref].value = formula
    print(f'{cell_ref}: {formula[:80]}{"..." if len(formula) > 80 else ""}')

wb.save('fetch-ibkr-positions-dashboard.xlsx')
print(f'\n✅ Updated {len(formulas)} formulas')
print('\n📝 Now open the file in Excel and click "Update Links" when prompted.')
