from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

wb = load_workbook('fetch-ibkr-positions-dashboard.xlsx')
ws = wb['Dashboard']

print('Fixing formula syntax...\n')

fixed_count = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.value:
            formula_text = None
            
            # Handle array formulas
            if isinstance(cell.value, ArrayFormula):
                formula_text = cell.value.text
            elif isinstance(cell.value, str) and cell.value.startswith('='):
                formula_text = cell.value
            
            if formula_text and "'[fetch-ibkr-positions.xlsx]'" in formula_text:
                # Fix the syntax: '[filename]'Sheet!Range -> '[filename]Sheet'!Range
                new_formula = formula_text.replace(
                    "'[fetch-ibkr-positions.xlsx]'PositionsHK!",
                    "'[fetch-ibkr-positions.xlsx]PositionsHK'!"
                ).replace(
                    "'[fetch-ibkr-positions.xlsx]'PositionsAL!",
                    "'[fetch-ibkr-positions.xlsx]PositionsAL'!"
                )
                
                if new_formula != formula_text:
                    cell.value = new_formula
                    print(f'{cell.coordinate}: Fixed')
                    fixed_count += 1

wb.save('fetch-ibkr-positions-dashboard.xlsx')
print(f'\n✅ Fixed {fixed_count} cells')
print('\nSample fixed formula:')
print(ws['E4'].value[:120] if ws['E4'].value else 'N/A')
